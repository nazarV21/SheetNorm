from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.ai.instruction_assistant import InstructionAssistant
from app.services.conversion_service import ConversionService


def _make_warehouse_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Исходные данные"
    rows = [
        ["Остатки материалов на складах", None, None, None, None, None],
        ["Дата: 2026-03-31", None, None, None, None, None],
        [None, None, None, None, None, None],
        ["Категория", "Материал", "Склад", "Количество", "Ед. изм.", "Примечание"],
        ["Трубы", "Труба 89 мм", "Склад 1", 58, "шт", None],
        [None, "Труба 89 мм", "Склад 2", 59, "шт", None],
        [None, "Труба 114 мм", "Склад 1", 230, "шт", None],
        [None, "Труба 114 мм", "Склад 2", 191, "шт", None],
        ["Реагенты", "Ингибитор коррозии", "Склад 1", 226, "л", None],
        [None, "Ингибитор коррозии", "Склад 2", 220, "л", None],
        [None, "Деэмульгатор", "Склад 1", 249, "л", None],
        [None, "Деэмульгатор", "Склад 2", 37, "л", None],
        ["Запчасти", "Фильтр", "Склад 1", 60, "шт", None],
        [None, "Фильтр", "Склад 2", 41, "шт", None],
        [None, "Клапан", "Склад 1", 216, "шт", None],
        [None, "Клапан", "Склад 2", 183, "шт", None],
    ]
    for row in rows:
        sheet.append(row)
    sheet.merge_cells("A1:F1")
    workbook.save(path)


def test_short_prompt_is_grounded_to_real_warehouse_column(app, tmp_path: Path):
    source = tmp_path / "warehouses.xlsx"
    _make_warehouse_workbook(source)

    with app.app_context():
        result = InstructionAssistant().prepare_instruction(
            source,
            "разделить по складам и создать для складов отдельные листы",
            sheet_name="Исходные данные",
        )

    assert result["analysis"]["header_rows"] == [3]
    assert result["analysis"]["data_start_row"] == 4
    assert result["analysis"]["table_type"] == "flat"
    assert result["user_intent"]["group_by_columns"] == ["Склад"]
    assert result["user_intent"]["fill_down_columns"] == ["Категория"]
    report = result["generated_rule"]["excel_report"]
    assert report["group_sheets"]["enabled"] is True
    assert report["group_sheets"]["group_by"] == ["Склад"]
    assert report["group_sheets"]["add_subtotal"] is False


def test_detailed_prompt_preserves_requested_columns_and_negations(app, tmp_path: Path):
    source = tmp_path / "warehouses.xlsx"
    _make_warehouse_workbook(source)
    prompt = (
        "Обработай лист «Исходные данные». Строка заголовков — 4, данные начинаются со строки 5. "
        "Заполни пустые значения в колонке «Категория» предыдущим непустым значением сверху. "
        "Раздели данные по уникальным значениям колонки «Склад». Для каждого склада создай отдельный лист "
        "и назови его значением из колонки «Склад». На каждом листе оставь колонки «Категория», «Материал», "
        "«Склад», «Количество», «Ед. изм.», «Примечание». Не создавай сводный или технический лист и не добавляй итоги."
    )

    with app.app_context():
        result = InstructionAssistant().prepare_instruction(source, prompt, sheet_name="Исходные данные")

    rule = result["generated_rule"]
    assert rule["header_rows"] == [3]
    assert rule["data_start_row"] == 4
    assert rule["select_columns"] == [
        "Категория",
        "Материал",
        "Склад",
        "Количество",
        "Ед. изм.",
        "Примечание",
    ]
    assert rule["fill_down_columns"] == ["Категория"]
    assert rule["excel_report"]["subtotal"]["enabled"] is False
    assert rule["excel_report"]["summary_sheet"]["enabled"] is False
    assert result["user_intent"]["no_totals"] is True
    assert result["user_intent"]["no_summary_sheet"] is True


def test_warehouse_conversion_creates_two_clean_sheets_without_totals(app, tmp_path: Path):
    source = tmp_path / "warehouses.xlsx"
    _make_warehouse_workbook(source)
    prompt = (
        "Обработай лист «Исходные данные». Строка заголовков — 4, данные начинаются со строки 5. "
        "Заполни пустые значения в колонке «Категория» предыдущим непустым значением сверху. "
        "Раздели данные по уникальным значениям колонки «Склад». Для каждого склада создай отдельный лист. "
        "На каждом листе оставь колонки «Категория», «Материал», «Склад», «Количество», «Ед. изм.», «Примечание». "
        "Не создавай сводный или технический лист и не добавляй итоги."
    )

    with app.app_context():
        assistant = InstructionAssistant().prepare_instruction(source, prompt, sheet_name="Исходные данные")
        rule = assistant["generated_rule"]
        service = ConversionService()
        dataframe = service._apply_generated_rule(source, rule)
        output = tmp_path / "result.xlsx"
        dataframe.to_excel(output, index=False, engine="openpyxl", sheet_name="Результат")
        service._format_excel_report(output, dataframe, {"generated_rule": rule}, source, {})

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["Склад 1", "Склад 2"]
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        assert [sheet.cell(1, col).value for col in range(1, 7)] == [
            "Категория",
            "Материал",
            "Склад",
            "Количество",
            "Ед. изм.",
            "Примечание",
        ]
        assert sheet.max_row == 7
        assert sheet.cell(2, 1).value == "Трубы"
        assert sheet.cell(3, 1).value == "Трубы"
        assert sheet.cell(4, 1).value == "Реагенты"
        assert sheet.cell(6, 1).value == "Запчасти"
        assert all(sheet.cell(row, 1).value != "Итого" for row in range(1, sheet.max_row + 1))
        assert sheet.auto_filter.ref == "A1:F7"
