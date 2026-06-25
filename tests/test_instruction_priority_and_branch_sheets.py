from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.ai.instruction_assistant import InstructionAssistant
from app.services.conversion_service import ConversionService


def _make_hierarchical_branch_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Исходные данные"
    rows = [
        ["Распределение затрат по объектам", None, None, None],
        ["Год: 2026", None, None, None],
        ["Филиал/Объект", "Статья затрат", "Сумма", "Валюта"],
        ["Филиал Север", None, None, None],
        ["Объект 1", "Материалы", 100, "RUB"],
        ["Объект 1", "Работы", 200, "RUB"],
        ["Филиал Юг", None, None, None],
        ["Объект 2", "Материалы", 300, "RUB"],
        ["Объект 2", "Работы", 400, "RUB"],
    ]
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_current_instruction_is_not_replaced_by_old_feedback(app, tmp_path: Path):
    source = tmp_path / "branches.xlsx"
    _make_hierarchical_branch_workbook(source)

    feedback_path = tmp_path / "instruction_feedback.json"
    feedback_path.write_text(
        json.dumps(
            [
                {
                    "id": "old-feedback",
                    "type": "instruction_revision",
                    "raw_prompt": "сделать фильтр по филиалу и периодам",
                    "previous_ai_instruction": "Перенести даты в строки и сформировать длинную таблицу.",
                    "user_corrected_instruction": "Добавить автофильтр, общий сводный лист и колонки Период и Показатель.",
                    "regenerated_instruction": "Создать общий лист Итоги и длинный формат.",
                    "generated_rule": {
                        "table_type": "cross_table",
                        "header_rows": [3, 4, 5],
                        "data_start_row": 6,
                        "melt": {"enabled": True},
                    },
                    "analysis_fingerprint": {
                        "table_type": "flat",
                        "header_depth": 1,
                        "has_total_rows": False,
                        "has_merged_cells": False,
                    },
                    "preview_columns": ["Филиал", "Период", "Показатель", "Значение"],
                }
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    app.config.update(
        INSTRUCTION_FEEDBACK_FILE=feedback_path,
        AI_USE_LEARNED_HINTS=True,
        AI_BACKEND="fallback",
    )
    prompt = "разделить по филиалам на разные листы и сделать итог для каждого"

    with app.app_context():
        result = InstructionAssistant().prepare_instruction(source, prompt)

    improved = result["ai_improved_instruction"].lower()
    assert "отдельный лист для каждого филиала" in improved
    assert "внизу каждого листа филиала" in improved
    assert "период" not in improved
    assert "длинн" not in improved
    assert "общий лист итоги" not in improved
    assert result["user_intent"]["needs_separate_sheets_by_branch"] is True
    assert result["user_intent"]["needs_per_sheet_totals"] is True
    assert result["user_intent"]["needs_summary_by_branch"] is False


def test_hierarchical_branches_are_written_to_separate_sheets_with_totals(app, tmp_path: Path):
    source = tmp_path / "branches.xlsx"
    _make_hierarchical_branch_workbook(source)
    app.config.update(
        INSTRUCTION_FEEDBACK_FILE=tmp_path / "feedback.json",
        AI_USE_LEARNED_HINTS=False,
        AI_BACKEND="fallback",
    )

    with app.app_context():
        assistant_result = InstructionAssistant().prepare_instruction(
            source,
            "разделить по филиалам на разные листы и сделать итог для каждого",
        )
        rule = assistant_result["generated_rule"]
        service = ConversionService()
        result_df = service._apply_generated_rule(source, rule)

        assert list(result_df.columns) == ["Филиал", "Объект", "Статья затрат", "Сумма", "Валюта"]
        assert result_df["Филиал"].tolist() == [
            "Филиал Север",
            "Филиал Север",
            "Филиал Юг",
            "Филиал Юг",
        ]

        output = tmp_path / "result.xlsx"
        result_df.to_excel(output, index=False, engine="openpyxl", sheet_name="Результат")
        service._format_excel_report(
            output,
            result_df,
            {"name": "Разовая инструкция", "generated_rule": rule},
            source,
            {
                "rows_input": 9,
                "rows_output": 4,
                "columns_input": 4,
                "columns_output": 5,
                "empty_cells_before": 0,
                "empty_cells_after": 0,
                "detected_table_type": "flat",
                "quality_status": "success",
                "confidence_score": 0.95,
                "warnings": [],
                "applied_operations": ["Определение филиалов", "Разделение по листам"],
            },
        )

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["Филиал Север", "Филиал Юг"]
    assert "Результат" not in workbook.sheetnames
    assert "Итоги" not in workbook.sheetnames

    north = workbook["Филиал Север"]
    south = workbook["Филиал Юг"]
    assert [north.cell(1, col).value for col in range(1, 5)] == ["Объект", "Статья затрат", "Сумма", "Валюта"]
    assert north["A2"].value == "Объект 1"
    assert north["C5"].value == "=SUBTOTAL(109,C2:C3)"
    assert south["C5"].value == "=SUBTOTAL(109,C2:C3)"
    assert north.auto_filter.ref == "A1:D3"
    assert south.auto_filter.ref == "A1:D3"


def test_repeated_llm_sentences_are_deduplicated():
    repeated = (
        "Создать отдельный лист для каждого филиала. "
        "Инструкция должна быть понятной аналитику. "
        "Инструкция должна быть понятной аналитику. "
        "Создать отдельный лист для каждого филиала."
    )
    cleaned = InstructionAssistant._clean_text(repeated)
    assert cleaned.count("Создать отдельный лист") == 1
    assert cleaned.count("Инструкция должна быть понятной аналитику") == 1


class _BadInstructionModel:
    def create_completion(self, **_kwargs):
        return {
            "choices": [
                {
                    "text": (
                        "Добавить автофильтр по филиалу, перенести периоды в строки, "
                        "создать общий лист Итоги и оставить показатели в длинном формате."
                    )
                }
            ]
        }


def test_llm_answer_that_ignores_current_request_is_rejected(app, tmp_path: Path):
    source = tmp_path / "branches.xlsx"
    _make_hierarchical_branch_workbook(source)
    app.config.update(
        INSTRUCTION_FEEDBACK_FILE=tmp_path / "feedback.json",
        AI_USE_LEARNED_HINTS=False,
        AI_BACKEND="fallback",
    )

    with app.app_context():
        assistant = InstructionAssistant()
        assistant._llm = _BadInstructionModel()
        assistant.backend = "llama_cpp"
        result = assistant.prepare_instruction(
            source,
            "разделить по филиалам на разные листы и сделать итог для каждого",
        )

    improved = result["ai_improved_instruction"].lower()
    assert "отдельный лист для каждого филиала" in improved
    assert "общий лист итоги" not in improved
    assert "период" not in improved
