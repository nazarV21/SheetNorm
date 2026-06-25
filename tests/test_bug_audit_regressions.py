from __future__ import annotations

import io
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from app.services.ai.llm_client import AIClient
from app.services.conversion_service import ConversionService, ConversionValidationError
from app.services.rule_schema import normalize_declarative_rule
from app.scripts_runtime.runner import ScriptExecutionError, ScriptRunner
from app.scripts_runtime.validator import validate_pandas_script
from app.utils.jobs_repository import JobsRepository
from app.utils.rules_store import RulesStore


def _workbook_bytes(rows: list[list[object]]) -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def test_calculated_columns_use_expression_and_accept_legacy_aliases(app, tmp_path: Path):
    source = tmp_path / "calc.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["a", "b"])
    sheet.append([2, 3])
    workbook.save(source)

    legacy_rule = {
        "table_type": "flat",
        "header_rows": [0],
        "data_start_row": 1,
        "calculated": [{"name": "total", "formula": "a + b"}],
    }
    normalized = normalize_declarative_rule(legacy_rule)
    assert normalized["calculated"] == [{"name": "total", "expression": "a + b"}]

    with app.app_context():
        stored = RulesStore().add_rule("calc", "calculate", generated_rule=legacy_rule)
        assert "formula" not in stored["generated_rule"]["calculated"][0]
        result = ConversionService()._apply_generated_rule(source, legacy_rule)
    assert result.loc[0, "total"] == 5


def test_invalid_calculated_expression_fails_instead_of_silent_success(app, tmp_path: Path):
    source = tmp_path / "bad-calc.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["a"])
    sheet.append([2])
    workbook.save(source)

    with app.app_context():
        with pytest.raises(ConversionValidationError) as caught:
            ConversionService()._apply_generated_rule(
                source,
                {
                    "table_type": "flat",
                    "header_rows": [0],
                    "data_start_row": 1,
                    "calculated": [{"name": "bad", "expression": "missing + 1"}],
                },
            )
    assert caught.value.diagnostics[0]["stage"] == "calculated"


def test_prompt_without_validated_rule_does_not_save_raw_success(app):
    with app.app_context():
        service = ConversionService()
        with pytest.raises(ConversionValidationError):
            service._convert_path_with_diagnostics(
                Path("unused.xlsx"),
                {"prompt": "rename columns", "use_raw_data": False},
                {},
                strict=False,
            )


def test_ai_client_no_longer_applies_transformations_directly(app):
    with app.app_context():
        with pytest.raises(RuntimeError):
            AIClient().apply_prompt(pd.DataFrame([{"a": 1}]), "return original")


def test_upload_rejects_html_file_with_xlsx_extension(client):
    response = client.post(
        "/api/uploads",
        data={"file": (io.BytesIO(b"<html><script>alert(1)</script></html>"), "evil.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 400
    assert response.get_json()["code"] == "INVALID_EXCEL_FILE"


def test_api_result_requires_success_and_safe_output_path(client, app, tmp_path: Path):
    with app.app_context():
        job = JobsRepository().create("failed-job", input_filename="in.xlsx", input_path=tmp_path / "in.xlsx")
        JobsRepository().update_status(job["job_id"], "failed", output_path=str(tmp_path / "secret.xlsx"), output_filename="secret.xlsx")
    response = client.get("/api/jobs/failed-job/result")
    assert response.status_code == 409
    assert response.get_json()["code"] == "RESULT_NOT_READY"


def test_web_download_requires_successful_job(client, app):
    output = Path(app.config["OUTPUT_DIR"]) / "orphan.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(_workbook_bytes([["a"], [1]]).getvalue())
    response = client.get("/download/orphan.xlsx")
    assert response.status_code == 404


def test_script_validator_blocks_introspection_and_bad_signature(app):
    assert not validate_pandas_script("import importlib\ndef transform(df):\n    return df\n").valid
    assert not validate_pandas_script("def transform(df):\n    return getattr(df, 'shape')\n").valid
    assert not validate_pandas_script("def transform(df, extra):\n    return df\n").valid
    assert not validate_pandas_script("def transform(df):\n    return df.__class__\n").valid

    with app.app_context():
        with pytest.raises(ScriptExecutionError):
            ScriptRunner.from_app_config().run("def transform(df):\n    return df.iloc[0:0]\n", pd.DataFrame([{"a": 1}]))

