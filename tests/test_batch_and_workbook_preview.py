from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook

from app.utils.jobs_repository import JobsRepository


def workbook_stream(*, sheets: tuple[str, ...] = ("Данные",)) -> io.BytesIO:
    workbook = Workbook()
    first = workbook.active
    first.title = sheets[0]
    first.append(["Филиал", "Сумма"])
    first.append(["Север", 10])
    for name in sheets[1:]:
        sheet = workbook.create_sheet(name)
        sheet.append(["Объект", "Значение"])
        sheet.append(["A", 20])
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def test_assistant_file_preview_is_available_before_analysis(client):
    response = client.post(
        "/api/assistant/file-preview",
        data={"file": (workbook_stream(sheets=("Исходные данные", "Справочник")), "report.xlsx")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert [sheet["name"] for sheet in payload["workbook_preview"]["sheets"]] == ["Исходные данные", "Справочник"]
    assert payload["prompt_tips"]
    assert payload["analysis"]["sheets"] == ["Исходные данные", "Справочник"]


def test_multiple_files_create_batch_page(client, monkeypatch):
    def fake_enqueue(job_id: str, **_kwargs):
        return {"job_id": job_id, "status": "queued"}

    monkeypatch.setattr("app.routes.web.enqueue_conversion", fake_enqueue)
    response = client.post(
        "/convert",
        data={
            "files": [
                (workbook_stream(), "north.xlsx"),
                (workbook_stream(), "south.xlsx"),
            ],
            "mode": "rule",
            "rule_id": "",
            "source_mode": "files",
        },
        content_type="multipart/form-data",
        follow_redirects=False,
    )

    assert response.status_code == 302
    assert "/batches/" in response.headers["Location"]
    page = client.get(response.headers["Location"])
    assert page.status_code == 200
    text = page.get_data(as_text=True)
    assert "north.xlsx" in text
    assert "south.xlsx" in text
    assert "Скачать готовые результаты ZIP" in text


def test_job_detail_renders_sheet_tabs_for_result(client, app):
    output_dir = Path(app.config["OUTPUT_DIR"])
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / "multi_result.xlsx"
    workbook = Workbook()
    workbook.active.title = "Филиал Север"
    workbook.active.append(["Объект", "Сумма"])
    workbook.active.append(["A", 10])
    south = workbook.create_sheet("Филиал Юг")
    south.append(["Объект", "Сумма"])
    south.append(["B", 20])
    workbook.save(output)

    source = Path(app.config["INPUT_DIR"]) / "source.xlsx"
    source.parent.mkdir(parents=True, exist_ok=True)
    source.write_bytes(b"placeholder")
    with app.app_context():
        repository = JobsRepository()
        repository.create("multi-preview", input_filename="source.xlsx", input_path=source)
        repository.update_result(
            "multi-preview",
            output_filename=output.name,
            output_path=output,
            quality_report={},
            duration_seconds=0.1,
        )

    response = client.get("/jobs/multi-preview")
    assert response.status_code == 200
    page = response.get_data(as_text=True)
    assert "Филиал Север" in page
    assert "Филиал Юг" in page
    assert "data-workbook-preview" in page


def test_assistant_preview_reads_workbook_without_dimension_metadata(client):
    import zipfile

    stream = workbook_stream(sheets=("Исходные данные",))
    source_bytes = stream.getvalue()
    source_zip = io.BytesIO(source_bytes)
    rebuilt = io.BytesIO()

    with zipfile.ZipFile(source_zip, "r") as archive, zipfile.ZipFile(rebuilt, "w", zipfile.ZIP_DEFLATED) as output:
        for item in archive.infolist():
            data = archive.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8")
                text = text.replace('<dimension ref="A1:B2"/>', "")
                data = text.encode("utf-8")
            output.writestr(item, data)
    rebuilt.seek(0)

    response = client.post(
        "/api/assistant/file-preview",
        data={"file": (rebuilt, "without-dimension.xlsx")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    payload = response.get_json()
    sheet = payload["workbook_preview"]["sheets"][0]
    assert sheet["total_rows"] == 2
    assert sheet["total_columns"] == 2
    assert len(sheet["rows"]) == 2
    assert payload["prompt_tips"]



def test_assistant_preview_ignores_incorrect_a1_dimension_metadata(client):
    import re
    import zipfile

    stream = workbook_stream(sheets=("Исходные данные",))
    source_zip = io.BytesIO(stream.getvalue())
    rebuilt = io.BytesIO()

    with zipfile.ZipFile(source_zip, "r") as archive, zipfile.ZipFile(rebuilt, "w", zipfile.ZIP_DEFLATED) as output:
        for item in archive.infolist():
            data = archive.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8")
                text = re.sub(r'<dimension ref="[^"]+"/>', '<dimension ref="A1:A1"/>', text)
                data = text.encode("utf-8")
            output.writestr(item, data)
    rebuilt.seek(0)

    response = client.post(
        "/api/assistant/file-preview",
        data={"file": (rebuilt, "wrong-dimension.xlsx")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    payload = response.get_json()
    sheet = payload["workbook_preview"]["sheets"][0]
    assert sheet["shown_rows"] == 2
    assert sheet["shown_columns"] == 2
    assert sheet["rows"][0]["cells"] == ["Филиал", "Сумма"]
    assert sheet["rows"][1]["cells"] == ["Север", "10"]

def test_validate_excel_file_closes_excel_handle(tmp_path, monkeypatch):
    from app.utils import uploads

    path = tmp_path / "book.xlsx"
    path.write_bytes(b"placeholder")
    state = {"closed": False}

    class FakeExcelFile:
        sheet_names = ["Данные"]

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            state["closed"] = True

    monkeypatch.setattr(uploads.pd, "ExcelFile", lambda _path: FakeExcelFile())

    valid, error = uploads.validate_excel_file(path)

    assert valid is True
    assert error == ""
    assert state["closed"] is True


def test_structure_analyzer_closes_openpyxl_workbook_before_pandas_read(tmp_path, monkeypatch):
    from app.services import table_structure_analyzer as analyzer_module

    path = tmp_path / "book.xlsx"
    path.write_bytes(workbook_stream().getvalue())
    real_load_workbook = analyzer_module.load_workbook
    real_read_excel = analyzer_module.pd.read_excel
    state = {"closed": False}

    class WorkbookProxy:
        def __init__(self, workbook):
            self._workbook = workbook

        @property
        def sheetnames(self):
            return self._workbook.sheetnames

        def __getitem__(self, key):
            return self._workbook[key]

        def close(self):
            state["closed"] = True
            self._workbook.close()

    monkeypatch.setattr(
        analyzer_module,
        "load_workbook",
        lambda *args, **kwargs: WorkbookProxy(real_load_workbook(*args, **kwargs)),
    )

    def checked_read_excel(*args, **kwargs):
        assert state["closed"] is True
        return real_read_excel(*args, **kwargs)

    monkeypatch.setattr(analyzer_module.pd, "read_excel", checked_read_excel)

    result = analyzer_module.TableStructureAnalyzer().analyze_excel(path)

    assert result["selected_sheet"] == "Данные"
    assert state["closed"] is True
