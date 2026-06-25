from __future__ import annotations

from pathlib import Path
from typing import Any
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


TOTAL_RE = re.compile(r"(^|\s)(итого|всего|total)(\s|$)", re.IGNORECASE)
HEADER_HINTS = {
    "объект", "филиал", "подразделение", "дата", "период", "сумма", "валюта",
    "наименование", "показатель", "значение", "статья затрат", "код",
}


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat(sep=" ")
        except TypeError:
            return value.isoformat()
    return str(value)


def _row_role(values: list[str], row_number: int) -> str:
    text = " ".join(value.strip() for value in values if value.strip())
    if TOTAL_RE.search(text):
        return "total"
    non_empty = [value.strip().lower() for value in values if value.strip()]
    if row_number == 1:
        return "header"
    if non_empty and sum(1 for value in non_empty if value in HEADER_HINTS) >= max(1, len(non_empty) // 2):
        return "header"
    return "data"


def build_workbook_preview(
    path: str | Path,
    *,
    max_rows: int = 40,
    max_columns: int = 18,
    exclude_sheets: set[str] | None = None,
) -> dict[str, Any]:
    """Return a compact, serializable workbook preview with sheet tabs.

    The function intentionally reads only a limited rectangle from every sheet.
    It is used both for Jinja rendering and for the assistant AJAX preview.
    """
    path = Path(path)
    excluded = {name.lower() for name in (exclude_sheets or set())}
    sheets: list[dict[str, Any]] = []
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception:
        workbook = None

    if workbook is not None:
        try:
            for worksheet in workbook.worksheets:
                if worksheet.title.lower() in excluded:
                    continue

                # Нельзя полностью доверять <dimension ref=...> внутри xlsx.
                # Некоторые сторонние системы записывают A1:A1, хотя на листе
                # фактически находятся десятки строк и колонок. В read_only
                # openpyxl тогда молча обрезает предпросмотр до одной ячейки.
                # Сохраняем заявленный размер, сбрасываем ограничение и читаем
                # только небольшой прямоугольник предпросмотра независимо от metadata.
                declared_rows = int(worksheet.max_row or 0)
                declared_columns = int(worksheet.max_column or 0)
                reset_dimensions = getattr(worksheet, "reset_dimensions", None)
                if callable(reset_dimensions):
                    reset_dimensions()

                scan_rows = max_rows + 1
                scan_columns = max_columns + 1
                raw_rows = list(
                    worksheet.iter_rows(
                        min_row=1,
                        max_row=scan_rows,
                        min_col=1,
                        max_col=scan_columns,
                        values_only=True,
                    )
                )

                last_non_empty_row = 0
                last_non_empty_column = 0
                for row_index, raw_row in enumerate(raw_rows, start=1):
                    row_last_column = 0
                    for column_index, value in enumerate(raw_row, start=1):
                        if value is not None and str(value).strip() != "":
                            row_last_column = column_index
                    if row_last_column:
                        last_non_empty_row = row_index
                        last_non_empty_column = max(last_non_empty_column, row_last_column)

                # Для полностью пустого, но существующего листа оставляем пустой
                # grid, а не рисуем фиктивную ячейку A1.
                shown_rows = min(last_non_empty_row, max_rows)
                shown_columns = min(last_non_empty_column, max_columns)
                if last_non_empty_row == 0 and last_non_empty_column == 0:
                    total_rows = 0
                    total_columns = 0
                else:
                    total_rows = max(declared_rows, last_non_empty_row)
                    total_columns = max(declared_columns, last_non_empty_column)
                has_more_rows = last_non_empty_row > max_rows
                has_more_columns = last_non_empty_column > max_columns

                columns = [get_column_letter(index) for index in range(1, shown_columns + 1)]
                rows: list[dict[str, Any]] = []
                for row_number, raw_row in enumerate(raw_rows[:shown_rows], start=1):
                    values = [
                        _display_value(value)
                        for value in raw_row[:shown_columns]
                    ]
                    rows.append(
                        {
                            "number": row_number,
                            "role": _row_role(values, row_number),
                            "cells": values,
                        }
                    )

                sheets.append(
                    {
                        "name": worksheet.title,
                        "columns": columns,
                        "rows": rows,
                        "total_rows": total_rows,
                        "total_columns": total_columns,
                        "shown_rows": shown_rows,
                        "shown_columns": shown_columns,
                        "is_limited": (
                            has_more_rows
                            or has_more_columns
                            or total_rows > shown_rows
                            or total_columns > shown_columns
                        ),
                    }
                )
        finally:
            workbook.close()
    else:
        # Legacy .xls files are not supported by openpyxl. Pandas/xlrd is used
        # as a read-only fallback when the corresponding optional dependency is available.
        with pd.ExcelFile(path) as excel:
            for sheet_name in excel.sheet_names:
                if sheet_name.lower() in excluded:
                    continue
                frame = pd.read_excel(excel, sheet_name=sheet_name, header=None, nrows=max_rows)
                frame = frame.where(pd.notna(frame), "")
                shown_rows = int(len(frame))
                shown_columns = min(int(len(frame.columns)), max_columns)
                rows = []
                for row_index, row in frame.iloc[:, :shown_columns].iterrows():
                    values = [_display_value(value) for value in row.tolist()]
                    rows.append({"number": int(row_index) + 1, "role": _row_role(values, int(row_index) + 1), "cells": values})
                sheets.append(
                    {
                        "name": sheet_name,
                        "columns": [get_column_letter(index) for index in range(1, shown_columns + 1)],
                        "rows": rows,
                        "total_rows": shown_rows,
                        "total_columns": shown_columns,
                        "shown_rows": shown_rows,
                        "shown_columns": shown_columns,
                        "is_limited": False,
                    }
                )
    return {
        "filename": path.name,
        "sheets": sheets,
        "sheet_count": len(sheets),
    }


def prompt_tips_from_analysis(analysis: dict[str, Any]) -> list[str]:
    """Build short, file-specific suggestions for the instruction textarea."""
    tips: list[str] = []
    sheets = analysis.get("sheets") or []
    selected_sheet = analysis.get("selected_sheet")
    header_rows = analysis.get("header_rows_human") or []
    data_start = analysis.get("data_start_row_human")
    fingerprint = analysis.get("fingerprint") or {}
    preview = analysis.get("preview") or []

    if len(sheets) > 1:
        tips.append(
            f"Укажите, какой лист обрабатывать. Сейчас выбран «{selected_sheet}», всего листов: {len(sheets)}."
        )
    elif selected_sheet:
        tips.append(f"Можно явно написать: «Обработать лист {selected_sheet}». ")

    if header_rows and data_start:
        human_headers = ", ".join(str(value) for value in header_rows)
        tips.append(f"Уточните структуру: заголовки на строке {human_headers}, данные начинаются со строки {data_start}.")

    if fingerprint.get("has_merged_cells"):
        tips.append("В файле есть объединённые ячейки. Укажите, нужно ли заполнить их значения вниз по строкам.")
    if fingerprint.get("has_total_rows"):
        tips.append("В файле найдены строки «Итого/Всего». Укажите, удалить их, сохранить или пересчитать заново.")
    if fingerprint.get("has_month_like_columns"):
        tips.append("Обнаружены месяцы в колонках. Укажите, нужно ли перенести их в отдельную колонку «Период».")

    first_column_values = []
    for row in preview[:30]:
        if row:
            value = str(row[0] or "").strip()
            if value:
                first_column_values.append(value.lower())
    if any(value.startswith(("филиал", "подразделение", "регион")) for value in first_column_values):
        tips.append("В первом столбце видны строки-разделители по филиалам. Укажите: общий лист или отдельный лист для каждого филиала, и нужны ли итоги.")

    tips.append("Перечислите колонки, которые должны остаться в результате, и задайте их желаемые названия.")
    return tips[:6]
