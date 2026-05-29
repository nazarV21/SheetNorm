from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


MONTH_PATTERNS = {
    "янв", "фев", "мар", "апр", "май", "июн", "июл", "авг", "сен", "сент", "окт", "ноя", "дек",
    "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec",
}
SERVICE_MARKERS = (
    "отчет", "отчёт", "организация", "период", "дата формирования", "дата отчета", "филиал:",
    "подразделение:", "ответственный", "примечание", "единица измерения", "форма", "лист",
)
TOTAL_MARKERS = ("итого", "всего", "total", "sum", "сумма")
HEADER_WORDS = ("заголов", "шапк", "назван", "колонк")
DATA_WORDS = ("данн", "таблиц", "строк", "запис")


@dataclass
class TablePreview:
    values: list[list[Any]]
    rows_count: int
    cols_count: int


class TableStructureAnalyzer:
    """Эвристический анализатор структуры Excel-файла.

    Важный принцип: явная инструкция пользователя имеет приоритет над
    автоматическими эвристиками. Если пользователь пишет «заголовки находятся
    на 4 строке» или «заголовки в одной строке», система обязана перестроить
    правило под это указание, а не продолжать считать таблицу многоуровневой.
    """

    def analyze_excel(
        self,
        path: str | Path,
        sheet_name: str | int | None = None,
        max_rows: int = 40,
        instruction_text: str | None = None,
    ) -> dict[str, Any]:
        path = Path(path)
        workbook = load_workbook(path, read_only=False, data_only=True)
        sheets = workbook.sheetnames
        selected_sheet = sheet_name if sheet_name not in (None, "") else sheets[0]
        if isinstance(selected_sheet, int):
            selected_sheet = sheets[selected_sheet]
        if selected_sheet not in sheets:
            selected_sheet = sheets[0]

        ws = workbook[selected_sheet]
        merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
        raw_df = pd.read_excel(path, sheet_name=selected_sheet, header=None, nrows=max_rows)
        raw_df = raw_df.dropna(how="all", axis=1)
        preview = self._preview(raw_df)

        non_empty_counts = [self._non_empty_count(row) for row in preview.values]
        first_data_like_row = self._find_first_data_like_row(preview.values, non_empty_counts)
        header_rows = self._guess_header_rows(preview.values, first_data_like_row)
        data_start_row = max(header_rows) + 1 if header_rows else first_data_like_row
        table_type = self._guess_table_type(preview.values, header_rows, data_start_row, merged_ranges)

        instruction_overrides = self._extract_instruction_overrides(instruction_text or "")
        header_rows, data_start_row, table_type = self._apply_instruction_overrides(
            rows=preview.values,
            header_rows=header_rows,
            data_start_row=data_start_row,
            table_type=table_type,
            first_data_like_row=first_data_like_row,
            overrides=instruction_overrides,
        )

        id_columns, value_columns = self._guess_id_and_value_columns(preview.values, header_rows, data_start_row)
        id_columns, value_columns = self._apply_column_overrides(id_columns, value_columns, instruction_overrides)
        service_rows = self._guess_service_rows(preview.values, data_start_row)
        total_rows = self._guess_total_rows(preview.values)

        fingerprint = {
            "sheet_count": len(sheets),
            "selected_sheet": selected_sheet,
            "rows_previewed": preview.rows_count,
            "cols_previewed": preview.cols_count,
            "merged_cells_count": len(merged_ranges),
            "has_merged_cells": bool(merged_ranges),
            "header_depth": len(header_rows),
            "table_type": table_type,
            "has_month_like_columns": self._has_month_like_columns(preview.values, header_rows),
            "has_total_rows": bool(total_rows),
            "starts_not_from_first_row": first_data_like_row > 0,
            "explicit_user_overrides": bool(instruction_overrides),
        }

        questions = self._build_questions(
            table_type,
            header_rows,
            id_columns,
            value_columns,
            service_rows,
            total_rows,
            instruction_overrides,
        )
        draft_instruction = self._build_draft_instruction(
            table_type=table_type,
            sheet_name=selected_sheet,
            header_rows=header_rows,
            data_start_row=data_start_row,
            id_columns=id_columns,
            value_columns=value_columns,
            service_rows=service_rows,
            total_rows=total_rows,
            instruction_overrides=instruction_overrides,
        )
        proposed_rule = self._build_rule(
            table_type=table_type,
            sheet_name=selected_sheet,
            header_rows=header_rows,
            data_start_row=data_start_row,
            id_columns=id_columns,
            value_columns=value_columns,
            total_rows=total_rows,
            instruction_overrides=instruction_overrides,
        )

        preview_context = self._build_preview_context(
            preview.values,
            header_rows=header_rows,
            data_start_row=data_start_row,
            service_rows=service_rows,
            total_rows=total_rows,
            max_rows=18,
            max_cols=14,
        )

        return {
            "file_name": path.name,
            "sheets": sheets,
            "selected_sheet": selected_sheet,
            "rows_previewed": preview.rows_count,
            "cols_previewed": preview.cols_count,
            "merged_ranges": merged_ranges[:30],
            "merged_ranges_count": len(merged_ranges),
            "first_data_like_row": first_data_like_row,
            "header_rows": header_rows,
            "data_start_row": data_start_row,
            "header_rows_human": [int(i) + 1 for i in header_rows],
            "data_start_row_human": int(data_start_row) + 1 if data_start_row is not None else None,
            "table_type": table_type,
            "id_columns": id_columns,
            "value_columns": value_columns,
            "service_rows": service_rows,
            "total_rows": total_rows,
            "fingerprint": fingerprint,
            "questions": questions,
            "draft_instruction": draft_instruction,
            "proposed_rule": proposed_rule,
            "preview": preview.values,
            "preview_context": preview_context,
            "instruction_overrides": instruction_overrides,
        }

    def _build_preview_context(
        self,
        rows: list[list[Any]],
        header_rows: list[int],
        data_start_row: int,
        service_rows: list[int],
        total_rows: list[int],
        max_rows: int = 18,
        max_cols: int = 14,
    ) -> dict[str, Any]:
        """Готовит предпросмотр с координатами и ролями строк для интерфейса."""
        col_count = max((len(row) for row in rows), default=0)
        col_count = min(col_count, max_cols)
        columns = [get_column_letter(i + 1) for i in range(col_count)]
        header_set = set(header_rows or [])
        service_set = set(service_rows or [])
        total_set = set(total_rows or [])

        prepared_rows: list[dict[str, Any]] = []
        for idx, row in enumerate(rows[:max_rows]):
            role = "data"
            role_label = "Данные"
            if idx in service_set:
                role = "service"
                role_label = "Служебная"
            if idx in header_set:
                role = "header"
                role_label = "Заголовок"
            if idx in total_set:
                role = "total"
                role_label = "Итог"
            if idx == data_start_row:
                role = "data-start"
                role_label = "Начало данных"

            cells = []
            for col_idx in range(col_count):
                value = row[col_idx] if col_idx < len(row) else ""
                cells.append({
                    "column": get_column_letter(col_idx + 1),
                    "value": value,
                    "is_empty": str(value).strip() == "",
                })
            prepared_rows.append({
                "number": idx + 1,
                "role": role,
                "role_label": role_label,
                "cells": cells,
            })

        legend = [
            {"role": "service", "label": "Служебные строки"},
            {"role": "header", "label": "Строки заголовков"},
            {"role": "data-start", "label": "Первая строка данных"},
            {"role": "total", "label": "Итоговые строки"},
        ]
        return {"columns": columns, "rows": prepared_rows, "legend": legend}

    def _preview(self, raw_df: pd.DataFrame) -> TablePreview:
        values: list[list[Any]] = []
        for _, row in raw_df.iterrows():
            cleaned = []
            for value in row.tolist():
                if pd.isna(value):
                    cleaned.append("")
                elif hasattr(value, "isoformat"):
                    cleaned.append(value.isoformat())
                else:
                    cleaned.append(str(value).strip())
            values.append(cleaned)
        rows_count = len(values)
        cols_count = max((len(row) for row in values), default=0)
        return TablePreview(values=values, rows_count=rows_count, cols_count=cols_count)

    @staticmethod
    def _non_empty_count(row: list[Any]) -> int:
        return sum(1 for cell in row if str(cell).strip() not in ("", "nan", "None"))

    def _find_first_data_like_row(self, rows: list[list[Any]], non_empty_counts: list[int]) -> int:
        if not rows:
            return 0
        max_count = max(non_empty_counts or [0])
        threshold = max(2, min(4, max_count // 2 if max_count else 2))
        for idx, row in enumerate(rows):
            text = " ".join(str(cell).lower() for cell in row if str(cell).strip())
            if non_empty_counts[idx] >= threshold and not self._is_service_text(text):
                return idx
        return 0

    def _guess_header_rows(self, rows: list[list[Any]], first_data_like_row: int) -> list[int]:
        """Определить заголовки без чрезмерного расширения на строки данных.

        Предыдущая логика часто считала 2–3 первые строки данных продолжением
        заголовка, если в них было мало чисел. Для типовых Excel-таблиц безопаснее
        считать заголовком одну строку и расширять его только при явных признаках
        многоуровневости: пустые ячейки в первой строке заголовка, периодические
        подписи, служебные строки над таблицей или отсутствие нормальных названий.
        """
        if not rows:
            return []
        start = max(0, min(first_data_like_row, len(rows) - 1))
        current = rows[start]
        next_row = rows[start + 1] if start + 1 < len(rows) else []

        if next_row and self._row_is_likely_data(next_row) and self._row_is_likely_header(current):
            return [start]

        # Если первая строка похожа на групповые заголовки, а следующая на имена колонок,
        # используем две строки. Иначе не растягиваем заголовок на данные.
        if next_row and self._row_has_many_blanks(current) and self._row_is_likely_header(next_row):
            return [start, start + 1]

        return [start]

    def _row_is_likely_header(self, row: list[Any]) -> bool:
        non_empty = [str(cell).strip() for cell in row if str(cell).strip()]
        if not non_empty:
            return False
        numeric = sum(1 for value in non_empty if self._looks_numeric(value))
        date_like = sum(1 for value in non_empty if self._looks_date(value))
        return (numeric + date_like) / len(non_empty) < 0.35

    def _row_is_likely_data(self, row: list[Any]) -> bool:
        non_empty = [str(cell).strip() for cell in row if str(cell).strip()]
        if not non_empty:
            return False
        numeric = sum(1 for value in non_empty if self._looks_numeric(value))
        date_like = sum(1 for value in non_empty if self._looks_date(value))
        # В строке продаж может быть только одна числовая колонка и дата, поэтому
        # не требуем большинства чисел.
        return (numeric + date_like) >= 1 and len(non_empty) >= 2

    def _row_has_many_blanks(self, row: list[Any]) -> bool:
        if not row:
            return False
        non_empty = self._non_empty_count(row)
        return non_empty <= max(2, len(row) // 3)

    def _guess_table_type(self, rows: list[list[Any]], header_rows: list[int], data_start_row: int, merged_ranges: list[str]) -> str:
        has_months = self._has_month_like_columns(rows, header_rows)
        # Не считаем таблицу кросс-таблицей только из-за большого числа числовых колонок:
        # обычные продажи тоже могут иметь количество, цену, сумму, скидку и т.д.
        if has_months:
            return "cross_table"
        if len(header_rows) > 1 or merged_ranges:
            return "multi_header"
        return "flat"

    def _guess_id_and_value_columns(self, rows: list[list[Any]], header_rows: list[int], data_start_row: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        if not rows:
            return [], []
        col_count = max((len(r) for r in rows), default=0)
        header_names = self._combined_headers(rows, header_rows, col_count)
        data_rows = rows[data_start_row : min(data_start_row + 12, len(rows))]
        id_columns: list[dict[str, Any]] = []
        value_columns: list[dict[str, Any]] = []

        for col_idx in range(col_count):
            values = [r[col_idx] if col_idx < len(r) else "" for r in data_rows]
            non_empty = [v for v in values if str(v).strip()]
            numeric_ratio = 0.0
            if non_empty:
                numeric_ratio = sum(1 for v in non_empty if self._looks_numeric(v)) / len(non_empty)
            header = header_names[col_idx] if col_idx < len(header_names) else f"Колонка {col_idx + 1}"
            item = {"index": col_idx, "name": header or f"Колонка {col_idx + 1}"}
            header_lower = str(header).lower()
            if numeric_ratio >= 0.55 or self._looks_period_header(header_lower):
                value_columns.append(item)
            elif str(header).strip() or non_empty:
                id_columns.append(item)

        if not value_columns and col_count > 2 and self._has_month_like_columns(rows, header_rows):
            id_columns = [{"index": i, "name": header_names[i] or f"Колонка {i + 1}"} for i in range(min(2, col_count))]
            value_columns = [{"index": i, "name": header_names[i] or f"Колонка {i + 1}"} for i in range(len(id_columns), col_count)]
        return id_columns[:8], value_columns[:30]

    def _combined_headers(self, rows: list[list[Any]], header_rows: list[int], col_count: int) -> list[str]:
        headers: list[str] = []
        for col_idx in range(col_count):
            parts: list[str] = []
            for row_idx in header_rows:
                if row_idx >= len(rows) or col_idx >= len(rows[row_idx]):
                    continue
                value = str(rows[row_idx][col_idx]).strip()
                if value:
                    parts.append(value)
            headers.append(" / ".join(dict.fromkeys(parts)) if parts else "")
        return headers

    def _has_month_like_columns(self, rows: list[list[Any]], header_rows: list[int]) -> bool:
        if not header_rows:
            return False
        checks = 0
        for row_idx in header_rows:
            if row_idx >= len(rows):
                continue
            for cell in rows[row_idx]:
                if self._looks_period_header(str(cell).lower()):
                    checks += 1
        return checks >= 2

    @staticmethod
    def _looks_numeric(value: Any) -> bool:
        text = str(value).strip().replace(" ", "").replace("%", "").replace(",", ".")
        if not text:
            return False
        try:
            float(text)
            return True
        except ValueError:
            return False

    @staticmethod
    def _looks_date(value: Any) -> bool:
        text = str(value).strip().lower()
        if not text:
            return False
        if re.search(r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b", text):
            return True
        if re.search(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b", text):
            return True
        return False

    @staticmethod
    def _looks_period_header(text: str) -> bool:
        text = str(text).lower().strip()
        if any(month in text for month in MONTH_PATTERNS):
            return True
        if re.search(r"\b20\d{2}\b", text):
            return True
        if re.search(r"\b[1-4]\s*(кв|квартал|q)\b", text):
            return True
        return False

    @staticmethod
    def _is_service_text(text: str) -> bool:
        text = text.lower()
        return any(marker in text for marker in SERVICE_MARKERS) and len(text.split()) <= 12

    def _guess_service_rows(self, rows: list[list[Any]], data_start_row: int) -> list[int]:
        result: list[int] = []
        for idx, row in enumerate(rows[:data_start_row]):
            text = " ".join(str(cell).lower() for cell in row if str(cell).strip())
            if text and (self._is_service_text(text) or self._non_empty_count(row) <= 2):
                result.append(idx)
        return result

    def _guess_total_rows(self, rows: list[list[Any]]) -> list[int]:
        result: list[int] = []
        for idx, row in enumerate(rows):
            text = " ".join(str(cell).lower() for cell in row if str(cell).strip())
            if any(marker in text for marker in TOTAL_MARKERS):
                result.append(idx)
        return result

    def _build_questions(
        self,
        table_type: str,
        header_rows: list[int],
        id_columns: list[dict[str, Any]],
        value_columns: list[dict[str, Any]],
        service_rows: list[int],
        total_rows: list[int],
        instruction_overrides: dict[str, Any] | None = None,
    ) -> list[str]:
        """Подсказки, которые можно добавить в инструкцию.

        Это не обязательные вопросы без ответа. Интерфейс показывает их как
        варианты уточнений, а пользователь может дописать ответ в инструкцию.
        """
        instruction_overrides = instruction_overrides or {}
        questions: list[str] = []
        if table_type == "cross_table":
            if not instruction_overrides.get("id_columns"):
                questions.append("Уточните в инструкции: какие колонки оставить как признаки объекта, например филиал, товар, дата, показатель.")
            if not instruction_overrides.get("melt_var_name"):
                questions.append("Уточните в инструкции: как назвать колонку для бывших заголовков периодов/групп, например 'Период'.")
            if not instruction_overrides.get("melt_value_name"):
                questions.append("Уточните в инструкции: как назвать колонку с числовыми значениями, например 'Продажи' или 'Значение'.")
        if len(header_rows) > 1 and not instruction_overrides.get("header_single_row"):
            questions.append("Если заголовки на самом деле в одной строке, напишите: 'заголовки находятся на строке N'.")
        if service_rows:
            questions.append("Если верхние служебные строки не нужны, напишите: 'служебные строки над таблицей удалить'.")
        if total_rows:
            questions.append("Если строки 'Итого'/'Всего' не должны попасть в результат, напишите: 'итоговые строки удалить'.")
        if not questions:
            questions.append("Если предпросмотр неверный, уточните в инструкции номер строки заголовков и строку начала данных.")
        return questions

    def _build_draft_instruction(
        self,
        table_type: str,
        sheet_name: str,
        header_rows: list[int],
        data_start_row: int,
        id_columns: list[dict[str, Any]],
        value_columns: list[dict[str, Any]],
        service_rows: list[int],
        total_rows: list[int],
        instruction_overrides: dict[str, Any] | None = None,
    ) -> str:
        instruction_overrides = instruction_overrides or {}
        human_header_rows = ", ".join(str(i + 1) for i in header_rows) if header_rows else "не определены"
        human_data_start = data_start_row + 1 if data_start_row is not None else "не определено"
        id_names = ", ".join(col["name"] for col in id_columns) or "уточнить у пользователя"
        value_names = ", ".join(col["name"] for col in value_columns[:12]) or "уточнить у пользователя"
        parts = [
            f"Обработать лист '{sheet_name}'.",
            f"Строки заголовков: {human_header_rows}; данные начинаются примерно со строки {human_data_start}.",
        ]
        if instruction_overrides:
            parts.append("Явные указания пользователя имеют приоритет над автоанализом файла.")
        if table_type == "cross_table":
            parts.append(
                "Таблица похожа на кросс-таблицу: нужно оставить идентификаторы "
                f"({id_names}), а колонки со значениями ({value_names}) преобразовать в строки."
            )
            parts.append("В итоговой таблице должны быть отдельные колонки для идентификаторов, периода/показателя и значения.")
        elif table_type == "multi_header":
            parts.append("Таблица содержит многоуровневый заголовок: нужно объединить уровни заголовков в понятные названия колонок.")
        else:
            parts.append("Таблица похожа на плоскую: нужно очистить заголовки, удалить лишние строки и привести данные к единому виду без лишнего расплавления колонок.")
        if service_rows:
            parts.append("Служебные строки перед основной таблицей удалить.")
        if total_rows:
            parts.append("Строки с итогами ('Итого', 'Всего') не включать в результат, если они не являются отдельными данными.")
        return " ".join(parts)

    def _build_rule(
        self,
        table_type: str,
        sheet_name: str,
        header_rows: list[int],
        data_start_row: int,
        id_columns: list[dict[str, Any]],
        value_columns: list[dict[str, Any]],
        total_rows: list[int],
        instruction_overrides: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        instruction_overrides = instruction_overrides or {}
        melt_var_name = instruction_overrides.get("melt_var_name") or "Период"
        melt_value_name = instruction_overrides.get("melt_value_name") or "Значение"
        melt_enabled = table_type == "cross_table" and not instruction_overrides.get("disable_melt")
        return {
            "table_type": table_type,
            "sheet_name": sheet_name,
            "header_rows": header_rows,
            "data_start_row": data_start_row,
            "id_columns": id_columns,
            "value_columns": value_columns,
            "drop_rows_contains": ["Итого", "Всего", "Total"] if total_rows else [],
            "normalize_headers": True,
            "instruction_overrides": instruction_overrides,
            "melt": {
                "enabled": melt_enabled,
                "var_name": melt_var_name,
                "value_name": melt_value_name,
            },
        }

    def _extract_instruction_overrides(self, text: str) -> dict[str, Any]:
        """Достать из инструкции явные указания пользователя.

        Поддерживаем формулировки вида:
        - «заголовки находятся на 4 строке»;
        - «заголовки в одной строке»;
        - «данные начинаются с 5 строки»;
        - «не расплавлять таблицу / не преобразовывать в длинный формат»;
        - «колонку периода назвать Месяц», «значения назвать Продажи».
        """
        normalized = self._normalize_instruction_text(text)
        if not normalized:
            return {}
        overrides: dict[str, Any] = {}
        header_sentences = [s for s in self._sentences(normalized) if any(w in s for w in ("заголов", "шапк"))]
        for sentence in header_sentences:
            if "одн" in sentence and "строк" in sentence:
                overrides["header_single_row"] = True
            row_range = self._extract_row_range(sentence)
            if row_range:
                overrides["header_rows"] = row_range
                overrides["header_single_row"] = len(row_range) == 1
                break
            row_number = self._extract_row_number(sentence)
            if row_number is not None:
                overrides["header_rows"] = [row_number]
                overrides["header_single_row"] = True
                break

        for sentence in self._sentences(normalized):
            if ("данн" in sentence or "таблиц" in sentence) and ("начина" in sentence or "старт" in sentence):
                row_number = self._extract_row_number(sentence)
                if row_number is not None:
                    overrides["data_start_row"] = row_number
                    break

        if any(phrase in normalized for phrase in ("не многоуров", "не несколько строк заголов", "заголовки не идут друг", "не объединять заголов")):
            overrides["header_single_row"] = True
        if any(phrase in normalized for phrase in ("не расплав", "не преобразовывать в длин", "не переводить в длин", "не melt", "не кросс")):
            overrides["table_type"] = "flat"
            overrides["disable_melt"] = True
        if any(phrase in normalized for phrase in ("кросс", "месяцы в колон", "периоды в колон", "преобразовать в строки", "перевести в длин")):
            overrides["table_type"] = "cross_table"

        id_cols = self._extract_named_list(normalized, ("идентификатор", "оставить как признаки", "оставить как идентификаторы", "признаки объекта"))
        if id_cols:
            overrides["id_columns"] = id_cols

        var_name = self._extract_output_column_name(normalized, ("период", "бывш", "заголов"))
        if var_name:
            overrides["melt_var_name"] = var_name
        value_name = self._extract_output_column_name(normalized, ("значен", "числов", "продаж", "сумм"))
        if value_name:
            overrides["melt_value_name"] = value_name
        return overrides

    @staticmethod
    def _normalize_instruction_text(text: str) -> str:
        text = (text or "").lower().replace("ё", "е")
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    @staticmethod
    def _sentences(text: str) -> list[str]:
        return [part.strip() for part in re.split(r"[.;!?\n]+", text) if part.strip()]

    def _extract_row_range(self, sentence: str) -> list[int] | None:
        match = re.search(r"строк[аиуыи]?\s*(?:№|номер)?\s*(\d+)\s*[-–—]\s*(\d+)", sentence)
        if not match:
            match = re.search(r"(?:с|со)\s*(\d+)\s*(?:по|до|[-–—])\s*(\d+)\s*строк", sentence)
        if not match:
            return None
        start = max(0, int(match.group(1)) - 1)
        end = max(0, int(match.group(2)) - 1)
        if start > end:
            start, end = end, start
        return list(range(start, end + 1))[:5]

    def _extract_row_number(self, sentence: str) -> int | None:
        patterns = [
            r"строк[аеуыи]?\s*(?:№|номер)?\s*(\d+)",
            r"(?:на|в|с|со)\s*(\d+)\s*(?:-?й|-?ой|-?ую|-?ю)?\s*строк",
        ]
        for pattern in patterns:
            match = re.search(pattern, sentence)
            if match:
                return max(0, int(match.group(1)) - 1)
        return None

    def _extract_named_list(self, text: str, markers: tuple[str, ...]) -> list[str]:
        for marker in markers:
            idx = text.find(marker)
            if idx < 0:
                continue
            fragment = text[idx: idx + 220]
            if ":" in fragment:
                fragment = fragment.split(":", 1)[1]
            else:
                fragment = re.sub(r"^.*?(?:это|будут|являются|оставить)\s+", "", fragment)
            fragment = re.split(r"[.;\n]", fragment)[0]
            parts = [self._clean_column_hint(p) for p in re.split(r",| и ", fragment)]
            parts = [p for p in parts if p and len(p) <= 50]
            if parts:
                return parts[:10]
        return []

    def _extract_output_column_name(self, text: str, markers: tuple[str, ...]) -> str | None:
        for marker in markers:
            if marker not in text:
                continue
            window_start = max(0, text.find(marker) - 80)
            window = text[window_start: text.find(marker) + 160]
            match = re.search(r"(?:назвать|назови|наименовать|имя|название)\s+[\"'«]?([a-zа-я0-9 _/-]{3,40})[\"'»]?", window)
            if match:
                return self._clean_column_hint(match.group(1)).title()
        return None

    @staticmethod
    def _clean_column_hint(text: str) -> str:
        text = re.sub(r"\b(колонк[ауе]?|пол[ея]?|как|нужно|надо|оставить|назвать|будут|будет)\b", "", text)
        text = re.sub(r"\s+", " ", text).strip(" ,.:;\"'«»")
        return text

    def _apply_instruction_overrides(
        self,
        *,
        rows: list[list[Any]],
        header_rows: list[int],
        data_start_row: int,
        table_type: str,
        first_data_like_row: int,
        overrides: dict[str, Any],
    ) -> tuple[list[int], int, str]:
        if not overrides:
            return header_rows, data_start_row, table_type
        if overrides.get("header_rows"):
            header_rows = [int(i) for i in overrides["header_rows"] if 0 <= int(i) < len(rows)] or header_rows
        elif overrides.get("header_single_row"):
            header_rows = [self._choose_best_single_header_row(rows, header_rows, first_data_like_row)]
        if overrides.get("header_single_row") and len(header_rows) > 1:
            header_rows = [self._choose_best_single_header_row(rows, header_rows, first_data_like_row)]
        if overrides.get("data_start_row") is not None:
            data_start_row = int(overrides["data_start_row"])
        else:
            data_start_row = max(header_rows) + 1 if header_rows else data_start_row
        if overrides.get("table_type"):
            table_type = str(overrides["table_type"])
        elif overrides.get("disable_melt"):
            table_type = "flat"
        return header_rows, data_start_row, table_type

    def _choose_best_single_header_row(self, rows: list[list[Any]], header_rows: list[int], first_data_like_row: int) -> int:
        candidates = header_rows or [first_data_like_row]
        candidates = [idx for idx in candidates if 0 <= idx < len(rows)] or [max(0, min(first_data_like_row, len(rows) - 1))]
        best_idx = candidates[0]
        best_score = -1.0
        for idx in candidates:
            row = rows[idx]
            non_empty = self._non_empty_count(row)
            numeric = sum(1 for cell in row if self._looks_numeric(cell) or self._looks_date(cell))
            text_score = non_empty - numeric
            score = text_score + (2 if self._row_is_likely_header(row) else 0)
            if score > best_score:
                best_idx = idx
                best_score = score
        return best_idx

    def _apply_column_overrides(
        self,
        id_columns: list[dict[str, Any]],
        value_columns: list[dict[str, Any]],
        overrides: dict[str, Any],
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        wanted_ids = overrides.get("id_columns") or []
        if not wanted_ids:
            return id_columns, value_columns
        all_cols = id_columns + value_columns
        selected: list[dict[str, Any]] = []
        for wanted in wanted_ids:
            wanted_lower = str(wanted).lower()
            for col in all_cols:
                name = str(col.get("name", ""))
                lower = name.lower()
                if wanted_lower and (wanted_lower == lower or wanted_lower in lower or lower in wanted_lower):
                    if col not in selected:
                        selected.append(col)
        if not selected:
            selected = [{"index": i, "name": name} for i, name in enumerate(wanted_ids)]
        selected_names = {str(col.get("name")) for col in selected}
        remaining_values = [col for col in all_cols if str(col.get("name")) not in selected_names and col in value_columns]
        return selected, remaining_values or value_columns
