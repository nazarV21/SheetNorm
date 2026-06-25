from __future__ import annotations

from pathlib import Path
from typing import Any
import threading
import json
import re
from datetime import datetime
from time import perf_counter
import tempfile

import pandas as pd
from flask import current_app
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.services.ai.llm_client import AIClient
from app.services.ai.smart_converter import SmartConverter
from app.utils.file_manager import FileManager
from app.utils.rules_store import RulesStore
from app.utils.training_examples_store import TrainingExamplesStore
from app.utils.cross_table_converter import convert_cross_table_to_flat
from app.utils.jobs_repository import JobsRepository
from app.services.rule_schema import normalize_declarative_rule
from app.services.workbook_preview import build_workbook_preview


class ConversionCancelledError(RuntimeError):
    """Raised when a user cancels a background conversion."""


class ConversionValidationError(Exception):
    """Ошибка, которую нужно показать пользователю на этапе предпросмотра/конвертации."""

    def __init__(self, message: str, diagnostics: list[dict[str, Any]] | None = None) -> None:
        super().__init__(message)
        self.message = message
        self.diagnostics = diagnostics or []


class ConversionService:
    """Основные шаги конвертации Excel-файла.

    Поддерживает три сценария:
    1. без правила — сохранить таблицу как есть;
    2. с сохранённым правилом/шаблоном;
    3. с разовой инструкцией пользователя, улучшенной AI-помощником.
    """

    def __init__(self, cancel_event: threading.Event | None = None) -> None:
        self.file_manager = FileManager(
            input_dir=current_app.config["INPUT_DIR"],
            output_dir=current_app.config["OUTPUT_DIR"],
        )
        self.ai_client = AIClient()
        self.smart_converter = SmartConverter()
        self.jobs = JobsRepository()
        self._job_started_at: dict[str, float] = {}
        self._cancel_event = cancel_event

    def _is_cancelled(self, job_id: str) -> bool:
        if self._cancel_event is not None and self._cancel_event.is_set():
            return True
        job = self.jobs.get(job_id)
        return bool(job and job.get("status") == "cancelled")

    def _ensure_active(self, job_id: str) -> None:
        if self._is_cancelled(job_id):
            raise ConversionCancelledError(f"Processing job {job_id} was cancelled.")

    def _begin_job(
        self,
        job_id: str,
        source_path: Path,
        rule: dict[str, Any] | None,
        instruction: str | None = None,
    ) -> None:
        self._ensure_active(job_id)
        if not self.jobs.get(job_id):
            self.jobs.create(
                job_id,
                input_filename=self.file_manager.get_original_filename(job_id),
                input_path=source_path,
            )
        self._job_started_at[job_id] = perf_counter()
        current = self.jobs.get(job_id) or {}
        self.jobs.update_status(
            job_id,
            "processing",
            stage="executing_transformation",
            progress=max(int(current.get("progress") or 0), 35),
            rule_id=(rule or {}).get("id"),
            rule_name=(rule or {}).get("name"),
            original_instruction=instruction,
            improved_instruction=(rule or {}).get("ai_improved_instruction") or instruction,
        )

    def _fail_job(
        self,
        job_id: str,
        error: str,
        *,
        code: str = "CONVERSION_FAILED",
        details: str | None = None,
    ) -> None:
        if self._is_cancelled(job_id):
            return
        self.jobs.append_error(job_id, error, code=code, details=details)

    @staticmethod
    def _validation_error_code(exc: ConversionValidationError) -> str:
        text = " ".join(
            [exc.message]
            + [str(item.get("message") or "") for item in exc.diagnostics]
            + [str(item.get("stage") or "") for item in exc.diagnostics]
        ).lower()
        if "лист" in text or "worksheet" in text or "sheet" in text:
            return "SHEET_NOT_FOUND"
        if "заголов" in text or "header" in text:
            return "HEADER_NOT_DETECTED"
        return "INVALID_RULE"

    def analyze(self, job_id: str) -> dict[str, Any]:
        try:
            sample_df = self._load_sample(job_id)
        except FileNotFoundError:
            return {"error": "Файл не найден", "job_id": job_id}
        suggestions = self.ai_client.suggest_schema(sample_df)
        return {"job_id": job_id, "suggestions": suggestions}

    def convert_with_rule(self, job_id: str, rule_id: str | None, options: dict[str, Any]) -> dict[str, Any]:
        source_path = self.file_manager.resolve_input(job_id)
        if not source_path.exists():
            self._fail_job(job_id, "Входной файл задачи не найден.", code="JOB_NOT_FOUND")
            return {"error": "Файл не найден", "job_id": job_id}

        rule = None
        if rule_id:
            rule = RulesStore().get_rule(rule_id)
            if not rule:
                self._fail_job(job_id, "Выбранный шаблон не найден.", code="INVALID_RULE")
                return {"error": "Выбранное правило не найдено", "job_id": job_id}

        self._begin_job(job_id, source_path, rule)
        try:
            result_df = self._convert_path(source_path=source_path, rule=rule, options=options)
        except ConversionValidationError as exc:
            code = self._validation_error_code(exc)
            self._fail_job(job_id, exc.message, code=code, details="Не удалось применить выбранный шаблон.")
            return {"error": exc.message, "details": "Не удалось применить выбранный шаблон.", "suggestion": "Проверьте структуру файла или создайте новый шаблон через AI-помощник.", "code": code, "diagnostics": exc.diagnostics}
        except Exception as exc:
            current_app.logger.exception("Rule conversion failed")
            self._fail_job(job_id, "Файл не удалось обработать", details=str(exc))
            return {"error": "Файл не удалось обработать", "details": str(exc), "suggestion": "Проверьте, что Excel-файл не повреждён и соответствует выбранному шаблону.", "code": "CONVERSION_FAILED"}
        self._ensure_active(job_id)
        return self._finish(job_id, source_path, result_df, rule)

    def _build_instruction_rule(
        self,
        instruction: str,
        generated_rule: dict[str, Any] | None,
        name: str = "Разовая инструкция",
    ) -> dict[str, Any]:
        return {
            "id": None,
            "name": name,
            "prompt": instruction,
            "ai_improved_instruction": instruction,
            "generated_rule": generated_rule or {},
            "use_raw_data": True,
            "sheet_name": (generated_rule or {}).get("sheet_name"),
        }

    def convert_with_instruction(
        self,
        job_id: str,
        instruction: str,
        generated_rule: dict[str, Any] | None = None,
        options: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Конвертация по разовой инструкции без обязательного сохранения правила."""
        source_path = self.file_manager.resolve_input(job_id)
        if not source_path.exists():
            self._fail_job(job_id, "Входной файл задачи не найден.", code="JOB_NOT_FOUND")
            return {"error": "Файл не найден", "job_id": job_id}
        rule = self._build_instruction_rule(instruction, generated_rule, name="Разовая инструкция")
        self._begin_job(job_id, source_path, rule, instruction)
        try:
            result_df = self._convert_path(source_path=source_path, rule=rule, options=options or {})
        except ConversionValidationError as exc:
            code = self._validation_error_code(exc)
            self._fail_job(job_id, exc.message, code=code, details="Инструкция не применилась к структуре файла.")
            return {"error": exc.message, "details": "Инструкция не применилась к структуре файла.", "suggestion": "Уточните строки заголовков, начало данных и требуемые колонки.", "code": code, "diagnostics": exc.diagnostics}
        except Exception as exc:
            current_app.logger.exception("Instruction conversion failed")
            self._fail_job(job_id, "Файл не удалось обработать", details=str(exc))
            return {"error": "Файл не удалось обработать", "details": str(exc), "suggestion": "Проверьте файл и уточните инструкцию.", "code": "CONVERSION_FAILED"}
        self._ensure_active(job_id)
        return self._finish(job_id, source_path, result_df, rule)

    def convert_with_instruction_checked(
        self,
        job_id: str,
        instruction: str,
        generated_rule: dict[str, Any] | None = None,
        options: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Финальная конвертация после предпросмотра.

        В отличие от старого режима, метод не молча возвращает сырой DataFrame при ошибке.
        Если внутреннее правило не применилось или результат выглядит подозрительно,
        пользователь получает понятное сообщение и должен уточнить инструкцию.
        """
        source_path = self.file_manager.resolve_input(job_id)
        if not source_path.exists():
            self._fail_job(job_id, "Входной файл задачи не найден.", code="JOB_NOT_FOUND")
            return {"error": "Файл не найден", "job_id": job_id, "diagnostics": []}
        rule = self._build_instruction_rule(instruction, generated_rule, name="Разовая инструкция")
        self._begin_job(job_id, source_path, rule, instruction)
        try:
            result_df, diagnostics = self._convert_path_with_diagnostics(
                source_path=source_path,
                rule=rule,
                options=options or {},
                strict=True,
            )
        except ConversionValidationError as exc:
            code = self._validation_error_code(exc)
            self._fail_job(job_id, exc.message, code=code)
            return {"error": exc.message, "job_id": job_id, "code": code, "diagnostics": exc.diagnostics}
        except Exception as exc:
            current_app.logger.exception("Checked conversion failed")
            self._fail_job(job_id, "Файл не удалось обработать", details=str(exc))
            return {"error": "Файл не удалось обработать", "job_id": job_id, "diagnostics": [{"severity": "error", "stage": "excel", "message": str(exc), "hint": "Проверьте целостность файла и выбранный лист."}]}
        self._ensure_active(job_id)
        result = self._finish(job_id, source_path, result_df, rule, diagnostics=diagnostics)
        result["diagnostics"] = diagnostics
        return result

    def preview_with_instruction(
        self,
        job_id: str,
        instruction: str,
        generated_rule: dict[str, Any] | None = None,
        max_rows: int = 100,
    ) -> dict[str, Any]:
        """Сформировать предпросмотр результата без финального скачивания.

        Если на этапе применения правила возникли ошибки, система не отдаёт
        случайный сырой результат, а возвращает диагностику для правки инструкции.
        """
        source_path = self.file_manager.resolve_input(job_id)
        if not source_path.exists():
            return {"error": "Файл не найден", "job_id": job_id, "diagnostics": []}
        rule = self._build_instruction_rule(instruction, generated_rule, name="Предпросмотр по инструкции")
        try:
            df, diagnostics = self._convert_path_with_diagnostics(
                source_path=source_path,
                rule=rule,
                options={},
                strict=True,
            )
        except ConversionValidationError as exc:
            return {"error": exc.message, "job_id": job_id, "diagnostics": exc.diagnostics}
        except Exception as exc:
            current_app.logger.exception("Preview conversion failed")
            return {"error": "Предпросмотр не сформирован", "job_id": job_id, "diagnostics": [{"severity": "error", "stage": "preview", "message": str(exc), "hint": "Проверьте целостность файла и уточните инструкцию."}]}
        preview = self.dataframe_to_preview(df, max_rows=max_rows)
        preview["total_rows"] = int(len(df))
        preview["total_columns"] = int(len(df.columns))
        return {
            "job_id": job_id,
            "source_filename": self.file_manager.get_original_filename(job_id),
            "dataframe": df,
            "preview": preview,
            "diagnostics": diagnostics,
        }

    def preview_workbook_with_instruction(
        self,
        job_id: str,
        instruction: str,
        generated_rule: dict[str, Any] | None = None,
        max_rows: int = 60,
        max_columns: int = 18,
    ) -> dict[str, Any]:
        """Build a sheet-aware preview that matches the final XLSX layout."""
        result = self.preview_with_instruction(
            job_id=job_id,
            instruction=instruction,
            generated_rule=generated_rule,
            max_rows=max_rows,
        )
        if "error" in result:
            return result

        dataframe = result.get("dataframe")
        if dataframe is None:
            return {
                "error": "Предпросмотр не сформирован",
                "job_id": job_id,
                "diagnostics": result.get("diagnostics") or [],
            }

        source_path = self.file_manager.resolve_input(job_id)
        rule = self._build_instruction_rule(instruction, generated_rule, name="Предпросмотр по инструкции")
        with tempfile.TemporaryDirectory(prefix="sheetnorm_preview_") as tmp:
            preview_path = Path(tmp) / "preview.xlsx"
            dataframe.to_excel(preview_path, index=False, engine="openpyxl", sheet_name="Результат")
            quality_report = {
                "rows_input": None,
                "rows_output": int(len(dataframe)),
                "columns_input": None,
                "columns_output": int(len(dataframe.columns)),
                "warnings": [],
                "applied_operations": [],
            }
            self._format_excel_report(preview_path, dataframe, rule, source_path, quality_report)
            workbook_preview = build_workbook_preview(
                preview_path,
                max_rows=max_rows,
                max_columns=max_columns,
            )
            original_name = self.file_manager.get_original_filename(job_id)
            workbook_preview["filename"] = f"{Path(original_name).stem}_preview.xlsx"

        return {
            "job_id": job_id,
            "source_filename": result.get("source_filename"),
            "workbook_preview": workbook_preview,
            "diagnostics": result.get("diagnostics") or [],
            "dataframe": dataframe,
        }

    def save_corrected_dataframe(
        self,
        job_id: str,
        df: pd.DataFrame,
        rule: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Сохранить пользовательски исправленную итоговую таблицу как Excel."""
        source_path = self.file_manager.resolve_input(job_id)
        if not source_path.exists():
            self._fail_job(job_id, "Входной файл задачи не найден.", code="JOB_NOT_FOUND")
            return {"error": "Файл не найден", "job_id": job_id}
        self._begin_job(job_id, source_path, rule)
        return self._finish(job_id, source_path, df, rule)

    @staticmethod
    def dataframe_to_preview(df: pd.DataFrame, max_rows: int = 100) -> dict[str, Any]:
        """Подготовить безопасные данные для HTML-предпросмотра итоговой таблицы."""
        limited = df.head(max_rows).copy()
        limited = limited.where(pd.notna(limited), "")
        columns = [str(col) for col in limited.columns]
        rows: list[list[str]] = []
        for _, row in limited.iterrows():
            rows.append([str(value) if value is not None else "" for value in row.tolist()])
        return {
            "columns": columns,
            "rows": rows,
            "shown_rows": len(rows),
            "is_limited": len(df) > max_rows,
        }


    @staticmethod
    def dataframe_to_editable_preview(df: pd.DataFrame, max_rows: int = 100) -> dict[str, Any]:
        """Совместимость со старым патчем: теперь это обычный предпросмотр без редактирования."""
        return ConversionService.dataframe_to_preview(df, max_rows=max_rows)

    @staticmethod
    def dataframe_from_editable_payload(payload: dict[str, Any]) -> pd.DataFrame:
        """Собрать DataFrame из таблицы, которую пользователь поправил в браузере."""
        columns = [str(col).strip() or f"Колонка {idx + 1}" for idx, col in enumerate(payload.get("columns") or [])]
        raw_rows = payload.get("rows") or []
        normalized_rows: list[list[Any]] = []
        for raw_row in raw_rows:
            row = list(raw_row) if isinstance(raw_row, list) else []
            if len(row) < len(columns):
                row.extend([""] * (len(columns) - len(row)))
            elif len(row) > len(columns):
                row = row[: len(columns)]
            # Не сохраняем полностью пустые строки, которые пользователь мог случайно добавить.
            if any(str(value).strip() for value in row):
                normalized_rows.append(row)
        return pd.DataFrame(normalized_rows, columns=columns)

    def convert(self, job_id: str, schema: dict | None, options: dict[str, Any]) -> dict[str, Any]:
        source_path = self.file_manager.resolve_input(job_id)
        if not source_path.exists():
            self._fail_job(job_id, "Входной файл задачи не найден.", code="JOB_NOT_FOUND")
            return {"error": "Файл не найден", "job_id": job_id}
        self._begin_job(job_id, source_path, None)
        try:
            df = pd.read_excel(source_path)
            if schema:
                df = self._apply_schema(df, schema)
        except Exception as exc:
            current_app.logger.exception("Schema conversion failed")
            self._fail_job(job_id, "Файл не удалось прочитать как Excel", details=str(exc))
            return {"error": "Файл не удалось прочитать как Excel", "details": str(exc), "suggestion": "Проверьте расширение и целостность файла.", "code": "INVALID_EXCEL_FILE"}
        self._ensure_active(job_id)
        return self._finish(job_id, source_path, df, None)

    def _convert_path(self, source_path: Path, rule: dict[str, Any] | None, options: dict[str, Any]) -> pd.DataFrame:
        df, _ = self._convert_path_with_diagnostics(source_path, rule, options, strict=False)
        return df

    def _convert_path_with_diagnostics(
        self,
        source_path: Path,
        rule: dict[str, Any] | None,
        options: dict[str, Any],
        strict: bool = False,
    ) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
        diagnostics: list[dict[str, Any]] = []

        def add(severity: str, stage: str, message: str, hint: str | None = None) -> None:
            item = {"severity": severity, "stage": stage, "message": message}
            if hint:
                item["hint"] = hint
            diagnostics.append(item)

        if not rule:
            try:
                df = pd.read_excel(source_path)
                diagnostics.extend(self._validate_converted_dataframe(df, stage="read_excel"))
                return df, diagnostics
            except Exception as exc:
                add("error", "read_excel", f"Не удалось прочитать Excel-файл: {exc}", "Проверьте, что файл не повреждён и имеет формат .xlsx/.xls.")
                raise ConversionValidationError("Файл не удалось прочитать как Excel.", diagnostics)

        prompt = rule.get("ai_improved_instruction") or rule.get("prompt") or ""
        generated_rule = normalize_declarative_rule(rule.get("generated_rule") or {})
        sheet_name = rule.get("sheet_name") or generated_rule.get("sheet_name") or 0
        use_raw_data = bool(rule.get("use_raw_data") or generated_rule)

        # 1. Формализованное правило имеет приоритет. В строгом режиме ошибка правила
        # не должна маскироваться возвратом сырой таблицы: пользователь должен увидеть,
        # что нужно уточнить инструкцию.
        if generated_rule:
            try:
                df_rule = self._apply_generated_rule(source_path, generated_rule)
                validation = self._validate_converted_dataframe(df_rule, stage="generated_rule")
                diagnostics.extend(validation)
                if any(item.get("severity") == "error" for item in validation):
                    raise ConversionValidationError(
                        "Результат применения инструкции выглядит некорректно. Уточните инструкцию и обновите предпросмотр.",
                        diagnostics,
                    )
                add("info", "generated_rule", "Внутреннее правило успешно применено к Excel-файлу.")
                return df_rule, diagnostics
            except ConversionValidationError:
                raise
            except Exception as exc:
                add(
                    "error",
                    "generated_rule",
                    f"Ошибка применения внутреннего правила: {exc}",
                    "Уточните в инструкции строку заголовков, строку начала данных, нужные колонки и необходимость преобразования в длинный формат.",
                )
                raise ConversionValidationError(
                    "Внутреннее правило не применилось. Исправьте инструкцию и обновите предпросмотр.",
                    diagnostics,
                )

        # 2. Базовое чтение Excel. Оно используется только как источник для few-shot/LLM,
        # но в строгом режиме не считается полноценным успешным результатом, если до этого
        # было сформировано правило и оно упало.
        try:
            df = pd.read_excel(source_path)
        except Exception as exc:
            add("error", "read_excel", f"Не удалось прочитать Excel-файл для fallback-этапа: {exc}")
            raise ConversionValidationError("Файл не удалось прочитать как Excel.", diagnostics) from exc

        if rule.get("id"):
            training_examples = TrainingExamplesStore().list_examples(rule_id=rule.get("id"))
        else:
            training_examples = TrainingExamplesStore().list_examples()

        if training_examples and prompt:
            try:
                if use_raw_data:
                    raw_df = self._read_raw(source_path, sheet_name)
                    df_smart = self.smart_converter.convert_with_examples(df, training_examples, prompt, raw_data=raw_df)
                else:
                    df_smart = self.smart_converter.convert_with_examples(df, training_examples, prompt)
                if df_smart is not None and len(df_smart.columns) > 0:
                    diagnostics.extend(self._validate_converted_dataframe(df_smart, stage="few_shot"))
                    add("info", "few_shot", "Результат построен с учётом обучающих примеров.")
                    return df_smart, diagnostics
            except Exception as exc:
                add("warning", "few_shot", f"Не удалось применить обучающие примеры: {exc}")

        if prompt:
            add(
                "error",
                "llm_prompt",
                "Текстовая инструкция не является исполняемым преобразованием без валидированного JSON-правила.",
                "Сформируйте preview через AI-помощник и сохраните правило перед финальной обработкой.",
            )
            raise ConversionValidationError(
                "Конвертация требует валидированного правила. Сырой результат не сохранён как успешный.",
                diagnostics,
            )

        if rule.get("type") == "cross_table":
            try:
                block_code = rule.get("block_code", "БВ-2.3а")
                raw_df = self._read_raw(source_path, sheet_name)
                df_cross = convert_cross_table_to_flat(raw_df, block_code=block_code)
                diagnostics.extend(self._validate_converted_dataframe(df_cross, stage="cross_table_fallback"))
                add("info", "cross_table_fallback", "Применён резервный алгоритм для кросс-таблиц.")
                return df_cross, diagnostics
            except Exception as exc:
                add("warning", "cross_table_fallback", f"Резервный алгоритм кросс-таблиц не сработал: {exc}")

        fallback_diagnostics = self._validate_converted_dataframe(df, stage="raw_fallback")
        diagnostics.extend(fallback_diagnostics)
        add(
            "warning",
            "raw_fallback",
            "Система не смогла уверенно применить инструкцию и получила результат обычным чтением Excel.",
            "Уточните инструкцию: где заголовки, где начинаются данные, какие строки удалить и какие итоги нужны.",
        )
        if strict:
            raise ConversionValidationError(
                "Конвертация требует уточнения инструкции. Система не должна отдавать сырой результат как готовый файл.",
                diagnostics,
            )
        return df, diagnostics

    def _validate_converted_dataframe(self, df: pd.DataFrame, stage: str) -> list[dict[str, Any]]:
        diagnostics: list[dict[str, Any]] = []
        if df is None:
            return [{"severity": "error", "stage": stage, "message": "Преобразование не вернуло таблицу."}]
        if len(df.columns) == 0:
            diagnostics.append({"severity": "error", "stage": stage, "message": "В результате нет колонок."})
        if len(df) == 0:
            diagnostics.append({"severity": "error", "stage": stage, "message": "В результате нет строк данных."})

        columns = [str(col).strip() for col in df.columns]
        generic = [col for col in columns if not col or col.lower().startswith("unnamed") or col.lower().startswith("колонка")]
        if columns and len(generic) / max(len(columns), 1) >= 0.5:
            diagnostics.append({
                "severity": "error",
                "stage": stage,
                "message": "Большая часть заголовков не распознана.",
                "hint": "Укажите в инструкции точную строку заголовков: например, 'заголовки находятся на строке 4, данные начинаются со строки 5'.",
            })
        elif generic:
            diagnostics.append({
                "severity": "warning",
                "stage": stage,
                "message": f"Есть нераспознанные заголовки: {', '.join(generic[:5])}.",
                "hint": "При необходимости уточните, какие колонки нужно оставить и как они должны называться.",
            })

        duplicates = sorted({col for col in columns if columns.count(col) > 1})
        if duplicates:
            diagnostics.append({
                "severity": "warning",
                "stage": stage,
                "message": f"Обнаружены повторяющиеся названия колонок: {', '.join(duplicates[:5])}.",
                "hint": "Уточните в инструкции, нужно ли объединять заголовки или переименовать одинаковые колонки.",
            })
        return diagnostics

    def _apply_generated_rule(self, source_path: Path, rule: dict[str, Any]) -> pd.DataFrame:
        rule = normalize_declarative_rule(rule or {})
        planning_error = rule.get("planning_error") or {}
        if planning_error:
            candidates = planning_error.get("candidates") or []
            hint = ""
            if candidates:
                hint = " Подходящие колонки: " + ", ".join(str(item) for item in candidates) + "."
            raise ConversionValidationError(
                str(planning_error.get("message") or "Не удалось построить план преобразования."),
                [{
                    "severity": "error",
                    "stage": "planning",
                    "message": str(planning_error.get("message") or "Не удалось определить колонку группировки."),
                    "hint": "Уточните колонку в инструкции, например: разделить по колонке «Склад»." + hint,
                }],
            )
        sheet_name = rule.get("sheet_name") or 0
        raw_df = self._read_raw(source_path, sheet_name)
        table_type = rule.get("table_type", "flat")
        header_rows = [int(i) for i in (rule.get("header_rows") or [0])]
        data_start_row = int(rule.get("data_start_row") if rule.get("data_start_row") is not None else (max(header_rows) + 1))
        drop_contains = [str(x).lower() for x in (rule.get("drop_rows_contains") or [])]

        col_count = raw_df.shape[1]
        headers = self._make_headers(raw_df, header_rows, col_count)
        data = raw_df.iloc[data_start_row:].copy()
        data = data.dropna(how="all")
        data.columns = headers[: data.shape[1]]
        data = data.reset_index(drop=True)
        data = self._normalize_column_names(data)

        selected_columns = [str(item).strip() for item in (rule.get("select_columns") or rule.get("columns") or []) if str(item).strip()]
        preserved_columns = self._match_existing_columns(data, selected_columns) if selected_columns else []

        if drop_contains:
            mask = pd.Series(False, index=data.index)
            for marker in drop_contains:
                row_text = data.astype(str).agg(" ".join, axis=1).str.lower()
                mask = mask | row_text.str.contains(marker, regex=False, na=False)
            data = data.loc[~mask].reset_index(drop=True)

        data = self._drop_empty_columns(data, preserve=preserved_columns)

        fill_down_columns = self._match_existing_columns(data, list(rule.get("fill_down_columns") or []))
        for column in fill_down_columns:
            data[column] = data[column].ffill()

        hierarchical_cfg = rule.get("hierarchical_groups") or {}
        if hierarchical_cfg.get("enabled"):
            data = self._expand_hierarchical_groups(data, hierarchical_cfg)

        if selected_columns:
            matched_selected = self._match_existing_columns(data, selected_columns)
            if not matched_selected:
                raise ConversionValidationError(
                    "Не удалось найти указанные колонки результата.",
                    [{
                        "severity": "error",
                        "stage": "select_columns",
                        "message": "Ни одна из запрошенных колонок не найдена в таблице.",
                        "hint": "Проверьте названия колонок в инструкции.",
                    }],
                )
            data = data[matched_selected].copy()

        for calculated_rule in rule.get("calculated") or []:
            target = calculated_rule.get("name")
            expression = calculated_rule.get("expression")
            if target and expression:
                try:
                    data[target] = data.eval(expression, engine="python")
                except Exception as exc:
                    raise ConversionValidationError(
                        f"Не удалось вычислить колонку '{target}'.",
                        [{
                            "severity": "error",
                            "stage": "calculated",
                            "message": str(exc),
                            "hint": "Проверьте expression и имена колонок.",
                        }],
                    ) from exc

        melt = rule.get("melt") or {}
        if table_type == "cross_table" or melt.get("enabled"):
            id_columns = [c.get("name") for c in (rule.get("id_columns") or []) if c.get("name")]
            value_columns = [c.get("name") for c in (rule.get("value_columns") or []) if c.get("name")]
            id_columns = [self._normalize_header(c) for c in id_columns if self._normalize_header(c) in data.columns]
            value_columns = [self._normalize_header(c) for c in value_columns if self._normalize_header(c) in data.columns]

            if not value_columns:
                id_set = set(id_columns)
                value_columns = [col for col in data.columns if col not in id_set and self._is_value_like_series(data[col])]
            if not id_columns:
                id_columns = [col for col in data.columns if col not in value_columns][:2]
            if value_columns:
                data = pd.melt(
                    data,
                    id_vars=id_columns,
                    value_vars=value_columns,
                    var_name=melt.get("var_name", "Период"),
                    value_name=melt.get("value_name", "Значение"),
                )
                data = data.dropna(subset=[melt.get("value_name", "Значение")], how="all")
                data = data.reset_index(drop=True)

        return data

    def _expand_hierarchical_groups(
        self,
        data: pd.DataFrame,
        config: dict[str, Any],
    ) -> pd.DataFrame:
        """Преобразовать строки-разделители вида «Филиал ...» в отдельную колонку.

        В исходных отчётах филиал часто хранится не в каждой строке, а отдельной
        строкой перед блоком объектов. Для разнесения по листам сначала нужно
        распространить значение филиала на все строки его блока.
        """
        if data.empty or len(data.columns) == 0:
            return data

        requested_source = str(config.get("source_column") or "").strip()
        matched = self._match_existing_columns(data, [requested_source]) if requested_source else []
        source_column = matched[0] if matched else data.columns[0]
        group_column = str(config.get("group_column_name") or "Филиал").strip() or "Филиал"
        detail_column = str(config.get("detail_column_name") or "Объект").strip() or str(source_column)
        marker_words = [str(item).lower() for item in (config.get("marker_words") or ["филиал", "подраздел", "регион"])]
        require_empty_other = bool(config.get("require_empty_other_cells", True))

        current_group: str | None = None
        keep_indexes: list[Any] = []
        groups: list[str | None] = []
        other_columns = [col for col in data.columns if col != source_column]

        for row_index, row in data.iterrows():
            raw_value = row.get(source_column)
            source_text = "" if pd.isna(raw_value) else str(raw_value).strip()
            other_has_values = any(
                not pd.isna(row.get(col)) and str(row.get(col)).strip() not in {"", "nan", "None"}
                for col in other_columns
            )
            is_marker = bool(source_text) and any(word in source_text.lower() for word in marker_words)
            if require_empty_other:
                is_marker = is_marker and not other_has_values

            if is_marker:
                current_group = source_text
                continue

            keep_indexes.append(row_index)
            groups.append(current_group)

        result = data.loc[keep_indexes].copy().reset_index(drop=True)
        result.insert(0, group_column, groups)
        if source_column in result.columns and source_column != detail_column:
            result = result.rename(columns={source_column: detail_column})
        if config.get("drop_ungrouped", False):
            result = result[result[group_column].notna()].reset_index(drop=True)
        return result

    def _read_raw(self, source_path: Path, sheet_name: str | int | None) -> pd.DataFrame:
        selected_sheet = sheet_name if sheet_name not in (None, "") else 0
        try:
            return pd.read_excel(source_path, sheet_name=selected_sheet, header=None)
        except ValueError as exc:
            if isinstance(selected_sheet, str):
                raise ConversionValidationError(
                    f"Лист '{selected_sheet}' не найден в Excel-файле.",
                    [{
                        "severity": "error",
                        "stage": "sheet",
                        "message": f"Не найден лист: {selected_sheet}.",
                        "hint": "Проверьте название листа в шаблоне или выберите существующий лист.",
                    }],
                ) from exc
            raise

    def _make_headers(self, raw_df: pd.DataFrame, header_rows: list[int], col_count: int) -> list[str]:
        headers: list[str] = []
        for col_idx in range(col_count):
            parts: list[str] = []
            for row_idx in header_rows:
                if row_idx >= raw_df.shape[0] or col_idx >= raw_df.shape[1]:
                    continue
                value = raw_df.iat[row_idx, col_idx]
                if pd.isna(value):
                    continue
                text = str(value).strip()
                if text and text.lower() != "nan":
                    parts.append(text)
            header = " / ".join(dict.fromkeys(parts)) if parts else f"Колонка {col_idx + 1}"
            headers.append(self._normalize_header(header))
        return headers

    @staticmethod
    def _normalize_header(text: Any) -> str:
        value = str(text).strip().replace("\n", " ")
        value = " ".join(value.split())
        return value or "Без названия"

    def _normalize_column_names(self, df: pd.DataFrame) -> pd.DataFrame:
        new_cols: list[str] = []
        seen: dict[str, int] = {}
        for col in df.columns:
            base = self._normalize_header(col)
            count = seen.get(base, 0)
            seen[base] = count + 1
            new_cols.append(base if count == 0 else f"{base}_{count + 1}")
        df.columns = new_cols
        return df

    @staticmethod
    def _drop_empty_columns(df: pd.DataFrame, preserve: list[str] | None = None) -> pd.DataFrame:
        preserve_set = {str(column) for column in (preserve or [])}
        removable = [
            column
            for column in df.columns
            if column not in preserve_set and df[column].isna().all()
        ]
        return df.drop(columns=removable) if removable else df

    @staticmethod
    def _is_value_like_series(series: pd.Series) -> bool:
        non_empty = series.dropna()
        if non_empty.empty:
            return False
        converted = pd.to_numeric(non_empty.astype(str).str.replace(" ", "", regex=False).str.replace(",", ".", regex=False), errors="coerce")
        return converted.notna().mean() >= 0.5

    def _load_sample(self, job_id: str, max_rows: int = 50) -> pd.DataFrame:
        path = self.file_manager.resolve_input(job_id)
        return pd.read_excel(path, nrows=max_rows)

    def _apply_schema(self, df: pd.DataFrame, schema: dict) -> pd.DataFrame:
        schema = normalize_declarative_rule(schema or {})
        rename_map = schema.get("rename", {})
        df = df.rename(columns=rename_map)
        select_cols = schema.get("columns") or schema.get("select_columns")
        if select_cols:
            df = df[[col for col in select_cols if col in df.columns]]
        calculated = schema.get("calculated", [])
        for rule in calculated:
            target = rule.get("name")
            expr = rule.get("expression")
            if target and expr:
                try:
                    df[target] = df.eval(expr)
                except Exception as exc:
                    raise ConversionValidationError(
                        f"Не удалось вычислить колонку '{target}'.",
                        [{
                            "severity": "error",
                            "stage": "calculated",
                            "message": str(exc),
                            "hint": "Проверьте expression и имена колонок.",
                        }],
                    ) from exc
        return df

    def _save_dataframe(
        self,
        df: pd.DataFrame,
        source_path: Path,
        job_id: str,
        rule: dict[str, Any] | None = None,
        quality_report: dict[str, Any] | None = None,
    ) -> Path:
        output_path = self.file_manager.prepare_output(job_id, extension=".xlsx")
        df.to_excel(output_path, index=False, engine="openpyxl", sheet_name="Результат")
        self._format_excel_report(output_path, df, rule, source_path, quality_report or {})
        return output_path

    def _format_excel_report(
        self,
        output_path: Path,
        df: pd.DataFrame,
        rule: dict[str, Any] | None,
        source_path: Path,
        quality_report: dict[str, Any],
    ) -> None:
        """Добавляет в итоговый Excel удобный отчётный слой.

        Если правило содержит excel_report, файл получает:
        - автофильтр по таблице, чтобы пользователь мог выбрать филиал;
        - закрепление строки заголовков;
        - итоговую строку с SUBTOTAL, которая пересчитывается после фильтра;
        - отдельный лист "Итоги" с агрегированными значениями по филиалам/подразделениям.
        """
        generated_rule = (rule or {}).get("generated_rule") or {}
        report = generated_rule.get("excel_report") or (rule or {}).get("excel_report") or {}

        try:
            wb = load_workbook(output_path)
            ws = wb.active
            ws.title = "Результат"
            max_row = ws.max_row
            max_col = ws.max_column
            if max_row < 1 or max_col < 1:
                wb.save(output_path)
                return

            header_fill = PatternFill("solid", fgColor="EAF2FF")
            total_fill = PatternFill("solid", fgColor="FFF7D6")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.freeze_panes = "A2" if report.get("freeze_header", True) else None

            data_last_row = max_row
            if report.get("auto_filter", True) and max_row >= 2:
                ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{data_last_row}"

            for idx, column_name in enumerate(df.columns, start=1):
                width = min(max(12, len(str(column_name)) + 2), 42)
                ws.column_dimensions[get_column_letter(idx)].width = width

            subtotal_cfg = report.get("subtotal") or {}
            numeric_columns = self._detect_numeric_columns(df)
            if subtotal_cfg.get("enabled") and numeric_columns and max_row >= 2:
                total_row = max_row + 2
                label_col = 1
                ws.cell(total_row, label_col).value = subtotal_cfg.get("label") or "Итого по выбранному фильтру"
                ws.cell(total_row, label_col).font = Font(bold=True)
                ws.cell(total_row, label_col).fill = total_fill
                for col_idx in numeric_columns:
                    col_letter = get_column_letter(col_idx + 1)
                    cell = ws.cell(total_row, col_idx + 1)
                    cell.value = f"=SUBTOTAL(109,{col_letter}2:{col_letter}{max_row})"
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
                for col_idx in range(1, max_col + 1):
                    ws.cell(total_row, col_idx).fill = total_fill

            group_sheets_cfg = report.get("group_sheets") or {}
            group_sheets_created = False
            if group_sheets_cfg.get("enabled"):
                group_sheets_created = self._add_group_sheets(wb, df, group_sheets_cfg)

            summary_cfg = report.get("summary_sheet") or {}
            if summary_cfg.get("enabled"):
                self._add_summary_sheet(wb, df, summary_cfg)

            if group_sheets_created and group_sheets_cfg.get("replace_result_sheet") and "Результат" in wb.sheetnames:
                del wb["Результат"]

            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.save(output_path)
        except Exception as exc:
            # Не скрываем причину ошибки отчётного слоя: основной результат уже
            # сохранён, но в журнале должно быть видно, почему листы/итоги не создались.
            current_app.logger.exception("Failed to format Excel report", exc_info=exc)
            return

    def _add_group_sheets(self, wb, df: pd.DataFrame, config: dict[str, Any]) -> bool:
        group_by = config.get("group_by") or ["Филиал"]
        group_columns = self._match_existing_columns(df, group_by)
        if not group_columns:
            current_app.logger.warning("Group sheets requested, but group column was not found: %s", group_by)
            return False

        group_column = group_columns[0]
        values = [value for value in df[group_column].dropna().drop_duplicates().tolist() if str(value).strip()]
        if not values:
            current_app.logger.warning("Group sheets requested, but no group values were found in column %s", group_column)
            return False

        created_names: set[str] = set(wb.sheetnames)
        created_count = 0
        for group_value in values:
            subset = df[df[group_column] == group_value].copy()
            if config.get("drop_group_column", True) and group_column in subset.columns:
                subset = subset.drop(columns=[group_column])
            if subset.empty:
                continue

            sheet_name = self._unique_sheet_name(str(group_value), created_names)
            created_names.add(sheet_name)
            ws = wb.create_sheet(sheet_name)
            for col_idx, col_name in enumerate(subset.columns, start=1):
                cell = ws.cell(1, col_idx, str(col_name))
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="EAF2FF")
                cell.alignment = Alignment(wrap_text=True, vertical="center")

            for row_idx, row in enumerate(subset.itertuples(index=False, name=None), start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row_idx, col_idx, self._excel_scalar(value))

            data_last_row = len(subset) + 1
            if data_last_row >= 2:
                ws.auto_filter.ref = f"A1:{get_column_letter(len(subset.columns))}{data_last_row}"
            ws.freeze_panes = "A2"

            if config.get("add_subtotal"):
                numeric_columns = self._detect_numeric_columns(subset)
                total_row = data_last_row + 2
                ws.cell(total_row, 1, config.get("subtotal_label") or "Итого")
                ws.cell(total_row, 1).font = Font(bold=True)
                total_fill = PatternFill("solid", fgColor="FFF7D6")
                for col_idx in range(1, len(subset.columns) + 1):
                    ws.cell(total_row, col_idx).fill = total_fill
                for zero_based_idx in numeric_columns:
                    col_idx = zero_based_idx + 1
                    col_letter = get_column_letter(col_idx)
                    cell = ws.cell(total_row, col_idx)
                    cell.value = f"=SUBTOTAL(109,{col_letter}2:{col_letter}{data_last_row})"
                    cell.font = Font(bold=True)

            for idx, column_name in enumerate(subset.columns, start=1):
                sample_lengths = [len(str(column_name))]
                sample_lengths.extend(len(str(value)) for value in subset[column_name].dropna().head(100).tolist())
                width = min(max(12, max(sample_lengths, default=12) + 2), 42)
                ws.column_dimensions[get_column_letter(idx)].width = width
            created_count += 1

        return created_count > 0

    @staticmethod
    def _unique_sheet_name(value: str, existing: set[str]) -> str:
        cleaned = re.sub(r"[\[\]:*?/\\]", "-", value).strip().strip("'") or "Группа"
        base = cleaned[:31]
        candidate = base
        suffix = 2
        while candidate in existing:
            ending = f" ({suffix})"
            candidate = base[: 31 - len(ending)] + ending
            suffix += 1
        return candidate

    @staticmethod
    def _excel_scalar(value: Any) -> Any:
        if value is None or (not isinstance(value, str) and pd.isna(value)):
            return None
        if hasattr(value, "item"):
            try:
                return value.item()
            except Exception:
                pass
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        return value

    def _add_summary_sheet(self, wb, df: pd.DataFrame, summary_cfg: dict[str, Any]) -> None:
        group_by = summary_cfg.get("group_by") or []
        group_columns = self._match_existing_columns(df, group_by)
        if not group_columns:
            group_columns = self._match_existing_columns(df, ["Филиал", "Подразделение", "Регион", "Город", "Branch", "Region"])
        numeric_indexes = self._detect_numeric_columns(df)
        numeric_columns = [df.columns[idx] for idx in numeric_indexes]
        if not numeric_columns:
            return

        work_df = df.copy()
        for col in numeric_columns:
            work_df[col] = pd.to_numeric(
                work_df[col].astype(str).str.replace(" ", "", regex=False).str.replace(",", ".", regex=False),
                errors="coerce",
            )
        if group_columns:
            summary = work_df.groupby(group_columns, dropna=False)[numeric_columns].sum(min_count=1).reset_index()
        else:
            summary = work_df[numeric_columns].sum(min_count=1).to_frame().T
            summary.insert(0, "Итог", "Всего")

        sheet_name = str(summary_cfg.get("sheet_name") or "Итоги")[:31]
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        for col_idx, col_name in enumerate(summary.columns, start=1):
            cell = ws.cell(1, col_idx, str(col_name))
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EAF2FF")
        for row_idx, row in enumerate(summary.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row_idx, col_idx, value)
        if summary.shape[0] >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(summary.shape[1])}{summary.shape[0] + 1}"
        ws.freeze_panes = "A2"
        for idx, column_name in enumerate(summary.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = min(max(14, len(str(column_name)) + 2), 42)

    def _detect_numeric_columns(self, df: pd.DataFrame) -> list[int]:
        numeric_columns: list[int] = []
        for idx, col in enumerate(df.columns):
            series = df[col].dropna()
            if series.empty:
                continue
            converted = pd.to_numeric(
                series.astype(str).str.replace(" ", "", regex=False).str.replace(",", ".", regex=False),
                errors="coerce",
            )
            if converted.notna().mean() >= 0.6:
                numeric_columns.append(idx)
        return numeric_columns

    def _match_existing_columns(self, df: pd.DataFrame, names: list[str]) -> list[str]:
        result: list[str] = []
        lowered = {str(col).lower(): col for col in df.columns}
        for wanted in names:
            wanted_lower = str(wanted).lower()
            for lower, original in lowered.items():
                if wanted_lower and (wanted_lower == lower or wanted_lower in lower or lower in wanted_lower):
                    if original not in result:
                        result.append(original)
        return result

    def _build_quality_report(
        self,
        source_path: Path,
        df: pd.DataFrame,
        rule: dict[str, Any] | None,
        diagnostics: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        metric_warnings = [str(item.get("message")) for item in (diagnostics or []) if item.get("severity") == "warning"]
        try:
            source_df = pd.read_excel(source_path, sheet_name=0, header=None)
            rows_input, columns_input = source_df.shape
            empty_before = int(source_df.isna().sum().sum())
        except Exception:
            rows_input, columns_input, empty_before = None, None, None
            metric_warnings.append("Не удалось рассчитать метрики исходного файла.")

        diagnostics = diagnostics or []
        warnings = metric_warnings
        errors = [str(item.get("message")) for item in diagnostics if item.get("severity") == "error"]
        generated_rule = (rule or {}).get("generated_rule") or {}
        operations: list[str] = []
        if generated_rule.get("header_rows"):
            operations.append("Определение заголовков")
        if generated_rule.get("drop_rows_contains"):
            operations.append("Удаление служебных/итоговых строк")
        if generated_rule.get("melt", {}).get("enabled") or generated_rule.get("table_type") == "cross_table":
            operations.append("Преобразование в длинный формат")
        if generated_rule.get("fill_down_columns"):
            operations.append("Заполнение иерархических значений сверху вниз")
        if (generated_rule.get("excel_report") or {}).get("group_sheets", {}).get("enabled"):
            operations.append("Разделение результата по отдельным листам")
        operations.append("Удаление пустых строк и колонок")

        quality_status = "error" if errors else ("warning" if warnings else "success")
        confidence = 0.75 if warnings else 0.92
        return {
            "rows_input": int(rows_input) if rows_input is not None else None,
            "rows_output": int(len(df)),
            "columns_input": int(columns_input) if columns_input is not None else None,
            "columns_output": int(len(df.columns)),
            "empty_cells_before": empty_before,
            "empty_cells_after": int(df.isna().sum().sum()),
            "warnings": warnings,
            "errors": errors,
            "detected_table_type": generated_rule.get("table_type") or (rule or {}).get("table_type") or "flat",
            "applied_operations": operations,
            "confidence_score": confidence,
            "quality_status": quality_status,
        }

    def _finish(
        self,
        job_id: str,
        source_path: Path,
        df: pd.DataFrame,
        rule: dict[str, Any] | None,
        diagnostics: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        self._ensure_active(job_id)
        current = self.jobs.get(job_id) or {}
        self.jobs.update_status(
            job_id,
            "processing",
            stage="validating_result",
            progress=max(int(current.get("progress") or 0), 75),
        )
        quality_report = self._build_quality_report(source_path, df, rule, diagnostics)
        self._ensure_active(job_id)
        self.jobs.update_status(job_id, "processing", stage="saving_result", progress=90)
        try:
            output_path = self._save_dataframe(df, source_path, job_id, rule=rule, quality_report=quality_report)
        except Exception as exc:
            current_app.logger.exception("Output save failed")
            self._fail_job(job_id, "Не удалось сохранить результат.", code="OUTPUT_SAVE_FAILED", details=str(exc))
            return {
                "error": "Не удалось сохранить результат",
                "details": str(exc),
                "suggestion": "Проверьте права на запись и свободное место в каталоге output.",
                "code": "OUTPUT_SAVE_FAILED",
                "job_id": job_id,
            }
        if self._is_cancelled(job_id):
            Path(output_path).unlink(missing_ok=True)
            raise ConversionCancelledError(f"Processing job {job_id} was cancelled before completion.")
        original_name = self.file_manager.get_original_filename(job_id)
        result = {
            "job_id": job_id,
            "source_filename": original_name,
            "rows": len(df),
            "columns": len(df.columns),
            "output": str(output_path),
            "output_filename": Path(output_path).name,
            "status": "converted",
            "rule_id": rule.get("id") if rule else None,
            "rule_name": rule.get("name") if rule else None,
            "quality_report": quality_report,
        }
        started_at = self._job_started_at.get(job_id)
        duration_seconds = perf_counter() - started_at if started_at is not None else None
        self.jobs.update_result(
            job_id,
            output_filename=Path(output_path).name,
            output_path=output_path,
            quality_report=quality_report,
            duration_seconds=duration_seconds,
            rule_id=(rule or {}).get("id"),
            rule_name=(rule or {}).get("name"),
            original_instruction=(rule or {}).get("raw_user_prompt") or (rule or {}).get("prompt"),
            improved_instruction=(rule or {}).get("ai_improved_instruction") or (rule or {}).get("prompt"),
        )
        self._append_history_entry(result)
        return result

    def _append_history_entry(self, result: dict[str, Any]) -> None:
        history_path = Path(current_app.config["HISTORY_FILE"])
        try:
            existing: list[dict[str, Any]] = json.loads(history_path.read_text(encoding="utf-8"))
        except FileNotFoundError:
            existing = []
        except json.JSONDecodeError:
            existing = []

        entry = {
            "job_id": result["job_id"],
            "source_filename": result.get("source_filename"),
            "output_filename": result.get("output_filename"),
            "rows": result.get("rows"),
            "columns": result.get("columns"),
            "status": result.get("status"),
            "rule_name": result.get("rule_name"),
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
        existing.append(entry)
        history_path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
